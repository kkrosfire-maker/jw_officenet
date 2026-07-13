"""
naver_output_2023 하위 폴더의 모든 Word 파일에서 텍스트를 추출하고
한국어 형태소 분석(kiwipiepy)으로 명사·동사·형용사 빈도를 집계해
기존 엑셀 파일에 새 탭으로 추가합니다.
"""
import os, re, sys
from collections import Counter
from docx import Document
from kiwipiepy import Kiwi
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

INPUT_ROOT = os.path.join(os.path.dirname(__file__), "naver_output_2023")
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "naver_2023_목록.xlsx")

SKIP_BRACKET = re.compile(r"^\[.+\]$")
SKIP_PREFIXES = ("원본 URL:", "작성일:", "http")

# 제외할 단어 목록 (불용어: 너무 일반적이거나 의미 없는 단어)
STOPWORDS = {
    "안녕하세요", "있습니다", "있는", "있어", "있고", "있으며", "있습니다",
    "합니다", "하는", "하고", "하여", "하면", "하지", "하게", "하기",
    "됩니다", "되는", "되어", "되고", "되면", "되지",
    "이런", "이렇게", "이러한", "이와", "이후", "이전",
    "그리고", "그런", "그래서", "그러나", "그러면", "그렇게", "그것",
    "때문에", "때문", "경우", "위해", "위한", "통해", "통한",
    "오늘", "우리", "저희", "여러", "다양한", "중요한", "중요",
    "관련", "필요", "가능", "정도", "이상", "이하", "특히",
    "수", "것", "등", "및", "의", "가", "이", "을", "를",
    "은", "는", "에", "에서", "로", "으로", "도", "과", "와",
    "이라", "라고", "이며", "이고", "지만",
    "많은", "많이", "또한", "함께", "통하여", "대한", "해당",
    "주요", "최근", "현재", "바로", "더욱", "매우", "너무",
    "좋은", "좋은", "좋습니다", "좋아", "같은", "같이",
    "연세예스내과", "내과", "인천", "주안", "미추홀", "주안동",
    "병원", "클리닉", "의원",
}

# Kiwi 품사 태그: NNG(일반명사), NNP(고유명사), VV(동사), VA(형용사)
TARGET_POS = {"NNG", "NNP", "VV", "VA"}


def extract_text_from_docx(path: str) -> str:
    try:
        doc = Document(path)
        lines = []
        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue
            if para.style.name.startswith("Heading"):
                continue
            if any(text.startswith(p) for p in SKIP_PREFIXES):
                continue
            if SKIP_BRACKET.match(text):
                continue
            lines.append(text)
        return " ".join(lines)
    except Exception:
        return ""


def main():
    # ── 1. 모든 Word 파일에서 텍스트 수집 ─────────────────────────────────────
    all_text_parts = []
    docx_count = 0
    for folder in sorted(os.listdir(INPUT_ROOT)):
        fp = os.path.join(INPUT_ROOT, folder)
        if not os.path.isdir(fp):
            continue
        for fname in os.listdir(fp):
            if fname.endswith(".docx"):
                text = extract_text_from_docx(os.path.join(fp, fname))
                if text:
                    all_text_parts.append(text)
                    docx_count += 1

    print(f"Word 파일 수: {docx_count}개")
    full_text = " ".join(all_text_parts)
    print(f"총 텍스트 길이: {len(full_text):,}자")

    # ── 2. 형태소 분석 ─────────────────────────────────────────────────────────
    print("형태소 분석 중...")
    kiwi = Kiwi()
    counter: Counter = Counter()

    # 청크 단위로 분석 (메모리 절약)
    chunk_size = 5000
    for i in range(0, len(full_text), chunk_size):
        chunk = full_text[i:i + chunk_size]
        result = kiwi.analyze(chunk)
        for token, pos, *_ in result[0][0]:
            if pos not in TARGET_POS:
                continue
            word = token.strip()
            if len(word) < 2:
                continue
            if word in STOPWORDS:
                continue
            if re.fullmatch(r"[0-9]+", word):
                continue
            counter[word] += 1

    top_words = counter.most_common(200)
    print(f"집계된 단어 종류: {len(counter):,}개 → 상위 {len(top_words)}개 저장")

    # ── 3. 기존 엑셀에 새 시트 추가 ───────────────────────────────────────────
    wb = openpyxl.load_workbook(EXCEL_PATH)

    sheet_name = "단어 빈도"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # 헤더 스타일
    header_font  = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="375623")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align   = Alignment(horizontal="center", vertical="center")
    num_align    = Alignment(horizontal="right",  vertical="center")
    thin         = Side(style="thin", color="BFBFBF")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["순위", "단어", "횟수"]
    col_widths = [8, 20, 12]

    ws.row_dimensions[1].height = 26
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    for rank, (word, count) in enumerate(top_words, 1):
        row = rank + 1
        ws.row_dimensions[row].height = 20

        ws.cell(row=row, column=1, value=rank).alignment  = cell_align
        ws.cell(row=row, column=2, value=word).alignment  = cell_align
        ws.cell(row=row, column=3, value=count).alignment = num_align

        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font   = Font(name="맑은 고딕", size=10)
            cell.border = border
            if rank % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EBF1DE")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{len(top_words)+1}"

    wb.save(EXCEL_PATH)
    print(f"\n저장 완료 -> {EXCEL_PATH}  (시트: '{sheet_name}')")

    # 상위 20개 미리보기
    print("\n상위 20개 단어:")
    print(f"{'순위':>4}  {'단어':<12}  {'횟수':>6}")
    print("-" * 28)
    for rank, (word, count) in enumerate(top_words[:20], 1):
        print(f"{rank:>4}  {word:<12}  {count:>6}")


if __name__ == "__main__":
    main()
