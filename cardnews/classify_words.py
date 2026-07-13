"""
단어 빈도 시트 D열에 분류(의료 / 지역·병원 / 기타)를 기입합니다.
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

EXCEL_PATH = r"C:\Users\JW\Desktop\workspace\cardnews\naver_2023_목록.xlsx"

# ── 분류 사전 ──────────────────────────────────────────────────────────────────
# 의료: 질병명, 증상, 신체부위, 검사·치료·약어, 건강 행위
MEDICAL = {
    "질환", "증상", "검사", "치료", "원인", "발생", "혈압", "내시경",
    "예방", "검진", "관리", "진료", "식도", "건강", "위암", "역류",
    "당뇨", "섭취", "대사", "발견", "대장", "증후군", "조기", "고지혈증",
    "수액", "운동", "위염", "비염", "만성", "부족", "심장", "염증",
    "혈관", "진단", "비만", "지방", "혈당", "기관지염", "혈액", "초기",
    "알레르기", "점막", "신체", "흡연", "내원", "스트레스", "약물",
    "인슐린", "골다공증", "당뇨병", "통증", "두통", "발병", "식습관",
    "비타민", "음주", "동맥", "감기", "축농증", "유발", "소화", "정기",
    "환자", "급성", "부정맥", "면역력", "복부", "불편", "체중",
    "소화기", "기능", "합병증", "뇌졸중", "콜레스테롤", "천식",
    "초음파", "저하", "정상", "수치", "전문의", "자극", "호흡",
    "소변", "신경", "가족력", "위험", "대장암", "분비", "건조",
    "부비동", "감염", "조절", "위산", "회복", "의심", "경화",
    "쓰리", "기침", "위장", "콧물", "바이러스", "심근", "경색",
    "방치", "복용", "주사", "카페인", "가슴", "부위",
}

# 지역·병원: 병원명, 지역명(시·구·동·로), 건물명
LOCATION_HOSPITAL = {
    "연세예스", "인천광역시", "미추홀구", "경인로", "주안동내과",
    "미추홀내과", "원장", "빌딩", "국제",
}

def classify(word: str) -> str:
    if word in MEDICAL:
        return "의료"
    if word in LOCATION_HOSPITAL:
        return "지역·병원"
    # 패턴 기반 추가 분류
    if any(word.endswith(s) for s in ("내과", "병원", "의원", "클리닉", "구", "시", "동", "로")):
        return "지역·병원"
    return "기타"


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["단어 빈도"]

    # ── D1 헤더 ───────────────────────────────────────────────────────────────
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    h = ws.cell(row=1, column=4, value="분류")
    h.font      = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    h.fill      = PatternFill("solid", fgColor="375623")
    h.alignment = Alignment(horizontal="center", vertical="center")
    h.border    = border
    ws.column_dimensions[get_column_letter(4)].width = 14

    # 색상 팔레트
    COLOR = {
        "의료":     "D9EAD3",  # 연두
        "지역·병원": "CFE2F3",  # 하늘
        "기타":     "F4CCCC",  # 연분홍
    }

    counts = {"의료": 0, "지역·병원": 0, "기타": 0}

    for row in ws.iter_rows(min_row=2):
        word_cell = row[1]  # B열
        word = word_cell.value
        if not word:
            continue

        label  = classify(word)
        counts[label] += 1
        r      = word_cell.row
        is_even = r % 2 == 0

        cell = ws.cell(row=r, column=4, value=label)
        cell.font      = Font(name="맑은 고딕", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        base_color = COLOR[label]

        # 짝수 행은 약간 어둡게 표시 (기존 행 색상과 혼합 효과 없이 단순 적용)
        cell.fill = PatternFill("solid", fgColor=base_color)

        # 짝수 행 기존 A~C 셀도 분류 색으로 통일
        if is_even:
            for col in range(1, 4):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=base_color)

    wb.save(EXCEL_PATH)

    print(f"저장 완료 -> {EXCEL_PATH}")
    print(f"\n분류 결과:")
    for label, cnt in counts.items():
        print(f"  {label:10s}: {cnt:3d}개")


if __name__ == "__main__":
    main()
