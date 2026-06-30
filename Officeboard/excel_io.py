"""엑셀 파싱/내보내기 로직. Flask 라우트와 완전히 분리."""
import re
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def _to_date_str(val) -> str:
    if val is None:
        return ''
    if isinstance(val, (date, datetime)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    return s.split(' ')[0] if ' ' in s else s


def parse_workbook(wb: openpyxl.Workbook) -> dict:
    """openpyxl Workbook → 대시보드 data dict 변환."""
    data = {}
    for ws in wb.worksheets:
        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
        if '호실' not in headers:
            continue

        # 납부 컬럼 및 세금계산서 컬럼 수집
        paid_cols = []
        invoice_cols = []
        for i, h in enumerate(headers):
            if '납부' in h:
                month_str = h.replace('납부', '').strip()
                paid_cols.append((i, month_str))
            elif '세금계산서' in h:
                month_str = h.replace('세금계산서', '').strip()
                invoice_cols.append((i, month_str))

        col = {h: i for i, h in enumerate(headers)}

        for row in ws.iter_rows(min_row=2, values_only=True):
            room_id = str(row[col.get('호실', 0)] or '').strip()
            if not room_id:
                continue

            name = str(row[col.get('상호명', col.get('입주자', 1))] or '').strip()
            tenant_name = str(row[col.get('입주자 이름', col.get('입주자', 2))] or '').strip()
            if not name and not tenant_name:
                continue

            rent_col = col.get('월세(원)')
            rent_val = row[rent_col] if rent_col is not None else None
            try:
                rent = int(float(str(rent_val))) if rent_val else 0
            except (ValueError, TypeError):
                rent = 0

            vat_col = col.get('VAT')
            vat_str = str(row[vat_col] or '') if vat_col is not None else ''
            vat = (vat_str.strip() == '유')

            disc_col = col.get('할인율')
            discount_str = str(row[disc_col] or '').strip() if disc_col is not None else ''
            if discount_str and discount_str != '표준가' and discount_str.endswith('%'):
                try:
                    discount = float(discount_str.replace('%', '')) / 100
                except (ValueError, TypeError):
                    discount = 0.0
            else:
                discount = 0.0

            contract_type = str(row[col.get('계약유형', 9)] or '입주').strip()
            if re.match(r'^V\d+$', room_id):
                contract_type = '비상주'

            entry = {
                'name': name,
                'tenantName': tenant_name,
                'phone': str(row[col.get('연락처', 3)] or '').strip(),
                'start': _to_date_str(row[col.get('계약시작', 4)]),
                'end': _to_date_str(row[col.get('계약종료', 5)]),
                'vat': vat,
                'discount': discount,
                'rent': rent,
                'contractType': contract_type,
                'memo': str(row[col.get('메모', 11)] or '').strip(),
            }

            if '계약상태' in col:
                cs = str(row[col['계약상태']] or '').strip()
                if cs in ('계약중', '계약만료', '계약해지'):
                    entry['contractStatus'] = cs

            if '선납처리월' in col:
                prepaid_at = str(row[col['선납처리월']] or '').strip()
                if prepaid_at:
                    entry['prepaidAt'] = prepaid_at

            # 모든 납부 컬럼 처리
            paid_months_all = []
            for paid_idx, paid_month in paid_cols:
                paid_val = row[paid_idx]
                entry['paid_' + paid_month] = (str(paid_val).strip() == '완납')
                paid_months_all.append(paid_month)

            # 세금계산서 컬럼 처리
            for inv_idx, inv_month in invoice_cols:
                inv_val = row[inv_idx]
                entry['invoice_' + inv_month] = (str(inv_val).strip() == '완료')

            # 계약기간 내 모든 월이 완납이면 prepaid 복원
            if paid_months_all and entry.get('start') and entry.get('end'):
                from datetime import date as _date
                contract_months = []
                s = entry['start'][:7]
                e = entry['end'][:7]
                cur_y, cur_m = int(s[:4]), int(s[5:7])
                end_y, end_m = int(e[:4]), int(e[5:7])
                while (cur_y, cur_m) <= (end_y, end_m):
                    contract_months.append(f'{cur_y}-{cur_m:02d}')
                    cur_m += 1
                    if cur_m > 12:
                        cur_m = 1
                        cur_y += 1
                if contract_months and all(entry.get('paid_' + m) for m in contract_months):
                    entry['prepaid'] = True

            data[room_id] = entry

    return data


def build_workbook(data: dict, month: str) -> openpyxl.Workbook:
    """data dict → 엑셀 Workbook 생성 (export 용)."""
    base_headers = [
        '호실', '상호명', '입주자 이름', '연락처',
        '계약시작', '계약종료', 'VAT', '할인율', '월세(원)', '계약유형', '선납처리월',
    ]
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='2C2C2C')
    header_align = Alignment(horizontal='center', vertical='center')

    def get_paid_months(room_ids):
        """해당 시트 호실들의 모든 paid_* 월을 수집해 정렬 반환."""
        months = set()
        for r in room_ids:
            d = data.get(r, {})
            for k in d:
                if k.startswith('paid_'):
                    months.add(k[5:])
        months.add(month)
        return sorted(months)

    def get_invoice_months(room_ids):
        """해당 시트 호실들의 모든 invoice_* 월을 수집해 정렬 반환."""
        months = set()
        for r in room_ids:
            d = data.get(r, {})
            for k in d:
                if k.startswith('invoice_'):
                    months.add(k[8:])
        months.add(month)
        return sorted(months)

    def write_sheet(ws, room_ids):
        paid_months    = get_paid_months(room_ids)
        invoice_months = get_invoice_months(room_ids)
        pay_headers     = [m + ' 납부' for m in paid_months]
        invoice_headers = [m + ' 세금계산서' for m in invoice_months]
        all_headers = base_headers + pay_headers + invoice_headers + ['메모', '계약상태']

        for col_idx, h in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 20

        for row_idx, room_id in enumerate(sorted(room_ids), 2):
            d = data[room_id]
            occupied = bool(d.get('name') or d.get('tenantName'))
            discount_pct = int(float(d.get('discount', 0)) * 100)
            discount_str = f'{discount_pct}%' if discount_pct else '표준가'

            pay_vals = []
            for m in paid_months:
                if occupied:
                    pk = 'paid_' + m
                    pay_vals.append('완납' if (d.get('prepaid') or d.get(pk)) else '미납')
                else:
                    pay_vals.append('')

            inv_vals = []
            for m in invoice_months:
                if occupied:
                    inv_vals.append('완료' if d.get('invoice_' + m) else '미발행')
                else:
                    inv_vals.append('')

            row_data = [
                room_id, d.get('name', ''), d.get('tenantName', ''),
                d.get('phone', ''), d.get('start', ''), d.get('end', ''),
                '유' if d.get('vat') else '무', discount_str,
                d.get('rent', 0) or '', d.get('contractType', ''),
                d.get('prepaidAt', '') or '',
            ] + pay_vals + inv_vals + [
                d.get('memo', ''),
                d.get('contractStatus', '계약중'),
            ]

            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        for col in ws.columns:
            width = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = width + 4

    resident_ids = [r for r in data if data[r].get('contractType') != '비상주']
    virtual_ids  = [r for r in data if data[r].get('contractType') == '비상주']

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = '상주'
    write_sheet(ws1, resident_ids)

    ws2 = wb.create_sheet('비상주')
    write_sheet(ws2, virtual_ids)
    return wb
