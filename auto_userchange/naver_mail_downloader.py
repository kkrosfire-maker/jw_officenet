import imaplib
import email
import email.header
import email.utils
import os
import re
import sys
import json
import base64
import threading
from dataclasses import dataclass, field
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import win32com.client as win32com
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False

try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except ImportError:
    HAS_DND = False


IMAP_HOST = "imap.naver.com"
IMAP_PORT = 993
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
CONFIG_DIR = os.path.join(os.environ.get("APPDATA", os.path.expanduser("~")), "NaverMailDownloader")
CONFIG_FILE = os.path.join(CONFIG_DIR, "config.json")
SOCKET_TIMEOUT = 60
HISTORY_MAX = 15
DEFAULT_REF_FILENAME = "정산파일담당자변형_기준엑셀.xlsx"


def get_default_ref_path() -> str:
    """실행 파일/스크립트 위치 기준으로 기본 기준 파일 경로 반환."""
    if getattr(sys, "frozen", False):
        # PyInstaller exe 실행 시
        base = os.path.dirname(sys.executable)
    else:
        # 일반 .py 실행 시
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, DEFAULT_REF_FILENAME)


# ── 유틸 ───────────────────────────────────────────────────────────────────────

def decode_header_value(value):
    if value is None:
        return ""
    parts = email.header.decode_header(value)
    result = []
    for part, charset in parts:
        if isinstance(part, bytes):
            result.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            result.append(part)
    return "".join(result)


def parse_date(date_str):
    if not date_str:
        return ""
    try:
        return email.utils.parsedate_to_datetime(date_str).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return date_str[:20]


def _encode_pw(pw):
    return base64.b64encode(pw.encode("utf-8")).decode("ascii")


def _decode_pw(encoded):
    return base64.b64decode(encoded.encode("ascii")).decode("utf-8")


def _is_ascii(text):
    return all(ord(c) < 128 for c in text)


def normalize_biz_no(value):
    """사업자번호를 숫자만 남겨서 정규화 (123-45-67890 → 1234567890)."""
    if value is None:
        return ""
    return re.sub(r"[^0-9]", "", str(value))


def parse_dnd_paths(data: str) -> list:
    """tkinterdnd2 drop 이벤트의 파일 경로 파싱 (공백 포함 경로 대응)."""
    paths = re.findall(r"\{([^}]+)\}|(\S+)", data)
    return [a or b for a, b in paths]


# ── 설정 파일 ──────────────────────────────────────────────────────────────────

def load_config() -> dict:
    try:
        with open(CONFIG_FILE, encoding="utf-8") as f:
            data = json.load(f)
        pw = _decode_pw(data["pw"]) if data.get("pw") else ""
        return {
            "user": data.get("user", ""),
            "pw": pw,
            "remember": data.get("remember", False),
            "sender_history": data.get("sender_history", []),
            "subject_history": data.get("subject_history", []),
            "ref_file_path": data.get("ref_file_path", ""),
        }
    except Exception:
        return {
            "user": "", "pw": "", "remember": False,
            "sender_history": [], "subject_history": [], "ref_file_path": "",
        }


def save_config(user, pw, remember, sender_history, subject_history, ref_file_path=""):
    os.makedirs(CONFIG_DIR, exist_ok=True)
    data = {
        "sender_history": sender_history,
        "subject_history": subject_history,
        "ref_file_path": ref_file_path,
    }
    if remember:
        data.update({"user": user, "pw": _encode_pw(pw), "remember": True})
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def clear_login_from_config():
    try:
        with open(CONFIG_FILE, encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        data = {}
    data.pop("user", None)
    data.pop("pw", None)
    data["remember"] = False
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


# ── 엑셀 변환 로직 ─────────────────────────────────────────────────────────────

@dataclass
class ReferenceData:
    """기준 엑셀 파일에서 로드한 매핑 데이터 전체."""
    mapping: dict = field(default_factory=dict)  # {정규화 사업자번호: 수탁업체명}
    names: list = field(default_factory=list)    # 수탁업체명 정렬 목록 (드롭다운용)
    clinic: dict = field(default_factory=dict)   # {품목명: 담당자} (윤병옥내과 시트)

    @classmethod
    def load(cls, path: str) -> "ReferenceData":
        """기준 파일을 한 번 열어 매핑 3종 모두 로드."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        ws0 = wb.worksheets[0]
        mapping: dict = {}
        name_set: set = set()
        for row in ws0.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            biz_no = normalize_biz_no(row[0])
            consignment = str(row[2]).strip() if len(row) >= 3 and row[2] is not None else ""
            if biz_no and consignment:
                mapping[biz_no] = consignment
            if consignment:
                name_set.add(consignment)

        clinic: dict = {}
        for ws in wb.worksheets:
            if ws.title == "윤병옥내과":
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    product = str(row[0]).strip()
                    consignment = str(row[1]).strip() if row[1] is not None else ""
                    if product and consignment:
                        clinic[product] = consignment
                break

        wb.close()
        return cls(mapping=mapping, names=sorted(name_set), clinic=clinic)


def _make_output_path(input_path: str, output_dir: str) -> str:
    today = datetime.now().strftime("%y%m%d")
    new_name = f"{today}_변환_{os.path.basename(input_path)}"
    output_path = os.path.join(output_dir, new_name)
    base, ext = os.path.splitext(output_path)
    counter = 1
    while os.path.exists(output_path):
        output_path = f"{base}_{counter}{ext}"
        counter += 1
    return output_path


def _unblock_file(path: str):
    """인터넷 출처 표시 제거 → Excel 보호된 보기 방지."""
    try:
        import subprocess
        subprocess.run(
            ["powershell", "-Command", f'Unblock-File -LiteralPath "{path}"'],
            capture_output=True, timeout=5,
        )
    except Exception:
        pass


class ExcelBackend:
    """Excel 파일 변환/셀 쓰기 백엔드 인터페이스."""

    def convert(self, input_path: str, mapping: dict, clinic_mapping: dict, output_dir: str) -> tuple:
        raise NotImplementedError

    def write_cells(self, path: str, rows: list) -> int:
        """rows: [(row_idx, new_value), ...] — 3번째 시트 B열(2열) 쓰기."""
        raise NotImplementedError


class Win32Backend(ExcelBackend):
    """Excel COM(win32com) 백엔드 — 슬라이서/차트/서식 완전 보존."""

    def convert(self, input_path: str, mapping: dict, clinic_mapping: dict, output_dir: str) -> tuple:
        import pythoncom
        pythoncom.CoInitialize()
        try:
            return self._convert_inner(input_path, mapping, clinic_mapping, output_dir)
        finally:
            pythoncom.CoUninitialize()

    def _convert_inner(self, input_path, mapping, clinic_mapping, output_dir) -> tuple:
        output_path = _make_output_path(input_path, output_dir)
        abs_in = os.path.abspath(input_path)
        abs_out = os.path.abspath(output_path)
        _unblock_file(abs_in)
        xl = win32com.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        xl.ScreenUpdating = False
        xl.EnableEvents = False
        xl.Interactive = False
        wb = None
        try:
            wb = xl.Workbooks.Open(abs_in, UpdateLinks=0, ReadOnly=False, AddToMru=False)
            ws = wb.Worksheets(3)
            last_row = ws.UsedRange.Rows.Count + ws.UsedRange.Row - 1
            changed, unmatched = 0, []
            for row in range(10, last_row + 1):
                hospital = ws.Cells(row, 6).Value
                if hospital and "윤병옥내과" in str(hospital).strip():
                    product = ws.Cells(row, 9).Value
                    key = str(product).strip() if product is not None else ""
                    if key and key in clinic_mapping:
                        ws.Cells(row, 2).Value = clinic_mapping[key]
                        changed += 1
                    elif key:
                        current_b = ws.Cells(row, 2).Value
                        unmatched.append({
                            "row": row,
                            "biz_raw": f"[윤병옥] {key}",
                            "current_b": str(current_b) if current_b is not None else "",
                        })
                    continue
                biz_raw = ws.Cells(row, 5).Value
                if biz_raw is None:
                    continue
                biz_no = normalize_biz_no(str(biz_raw))
                if not biz_no:
                    continue
                if biz_no in mapping:
                    ws.Cells(row, 2).Value = mapping[biz_no]
                    changed += 1
                else:
                    current_b = ws.Cells(row, 2).Value
                    unmatched.append({
                        "row": row,
                        "biz_raw": str(biz_raw),
                        "current_b": str(current_b) if current_b is not None else "",
                    })
            fmt = 52 if abs_in.lower().endswith(".xlsm") else 51
            wb.SaveAs(abs_out, FileFormat=fmt)
        finally:
            if wb is not None:
                try:
                    wb.Close(False)
                except Exception:
                    pass
            xl.Quit()
        return output_path, changed, unmatched

    def write_cells(self, path: str, rows: list) -> int:
        import pythoncom
        pythoncom.CoInitialize()
        try:
            abs_path = os.path.abspath(path)
            xl = win32com.DispatchEx("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False
            xl.ScreenUpdating = False
            xl.EnableEvents = False
            xl.Interactive = False
            wb = None
            try:
                wb = xl.Workbooks.Open(abs_path, UpdateLinks=0, ReadOnly=False, AddToMru=False)
                ws = wb.Worksheets(3)
                for row_idx, new_val in rows:
                    ws.Cells(row_idx, 2).Value = new_val
                wb.Save()
            finally:
                if wb is not None:
                    try:
                        wb.Close(False)
                    except Exception:
                        pass
                xl.Quit()
        finally:
            pythoncom.CoUninitialize()
        return len(rows)


class OpenpyxlBackend(ExcelBackend):
    """openpyxl 백엔드 — 슬라이서가 제거될 수 있음."""

    def convert(self, input_path: str, mapping: dict, clinic_mapping: dict, output_dir: str) -> tuple:
        wb = openpyxl.load_workbook(input_path)
        ws = wb.worksheets[2]
        changed, unmatched = 0, []
        for row_idx in range(10, ws.max_row + 1):
            hospital = ws.cell(row=row_idx, column=6).value
            if hospital and "윤병옥내과" in str(hospital).strip():
                product = ws.cell(row=row_idx, column=9).value
                key = str(product).strip() if product is not None else ""
                if key and key in clinic_mapping:
                    ws.cell(row=row_idx, column=2).value = clinic_mapping[key]
                    changed += 1
                elif key:
                    current_b = ws.cell(row=row_idx, column=2).value
                    unmatched.append({
                        "row": row_idx,
                        "biz_raw": f"[윤병옥] {key}",
                        "current_b": str(current_b) if current_b is not None else "",
                    })
                continue
            biz_raw = ws.cell(row=row_idx, column=5).value
            if biz_raw is None:
                continue
            biz_no = normalize_biz_no(biz_raw)
            if not biz_no:
                continue
            if biz_no in mapping:
                ws.cell(row=row_idx, column=2).value = mapping[biz_no]
                changed += 1
            else:
                current_b = ws.cell(row=row_idx, column=2).value
                unmatched.append({
                    "row": row_idx,
                    "biz_raw": str(biz_raw),
                    "current_b": str(current_b) if current_b is not None else "",
                })
        output_path = _make_output_path(input_path, output_dir)
        wb.save(output_path)
        return output_path, changed, unmatched

    def write_cells(self, path: str, rows: list) -> int:
        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[2]
        for row_idx, new_val in rows:
            ws.cell(row=row_idx, column=2).value = new_val
        wb.save(path)
        return len(rows)


def get_excel_backend() -> ExcelBackend:
    if HAS_WIN32:
        return Win32Backend()
    elif HAS_OPENPYXL:
        return OpenpyxlBackend()
    else:
        raise RuntimeError("pywin32 또는 openpyxl 패키지가 필요합니다.")


def convert_file(input_path: str, mapping: dict, clinic_mapping: dict, output_dir: str) -> tuple:
    """슬라이서 등 Excel 요소를 보존하며 변환. 반환: (출력 경로, 변환된 행 수, 미매칭 목록)."""
    return get_excel_backend().convert(input_path, mapping, clinic_mapping, output_dir)


def _append_to_reference(ref_path: str, ref_additions: list):
    """기준 파일에 신규 매핑 추가. ref_additions: [(biz_raw_or_clinic_key, 수탁업체명)]"""
    wb = openpyxl.load_workbook(ref_path)

    normal = [(k, v) for k, v in ref_additions if not k.startswith("[윤병옥]")]
    clinic = [(k[len("[윤병옥] "):], v) for k, v in ref_additions if k.startswith("[윤병옥]")]

    if normal:
        ws0 = wb.worksheets[0]
        existing = set()
        for row in ws0.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                existing.add(normalize_biz_no(row[0]))
        for biz_raw, consignment in normal:
            norm = normalize_biz_no(biz_raw)
            if norm and norm not in existing:
                ws0.append([biz_raw, None, consignment])
                existing.add(norm)

    if clinic:
        for ws in wb.worksheets:
            if ws.title == "윤병옥내과":
                existing = {str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]}
                for product, consignment in clinic:
                    if product and product not in existing:
                        ws.append([product, consignment])
                        existing.add(product)
                break

    wb.save(ref_path)


# ── 앱 ────────────────────────────────────────────────────────────────────────

class NaverMailApp:
    def __init__(self, root):
        self.root = root
        self.root.title("네이버 메일 첨부파일 다운로더")
        self.root.resizable(True, True)

        # IMAP
        self.imap = None
        self.mail_list = []
        self._saved_user = ""
        self._saved_pw = ""
        self._remember = False
        self._sender_history = []
        self._subject_history = []
        self._imap_lock = threading.Lock()

        # 변환
        self._ref_file_path = ""
        self._ref_data: ReferenceData = ReferenceData()
        self._unmatched_data: dict = {}
        self._edit_entry = None

        self._build_ui()
        self.root.update_idletasks()
        self.root.geometry(f"940x{self.root.winfo_reqheight()}")
        self._load_config_into_ui()

    # ── UI 최상위 ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        self._build_status_bar()          # bottom 먼저 pack
        self._build_login_frame()
        self._build_search_frame()
        self._build_result_frame()
        self._build_convert_section()
        self._build_unmatched_section()

    # ── 로그인 ────────────────────────────────────────────────────────────────

    def _build_login_frame(self):
        frame = tk.LabelFrame(self.root, text="로그인", padx=10, pady=6)
        frame.pack(fill="x", padx=10, pady=(8, 3))

        tk.Label(frame, text="네이버 아이디:").grid(row=0, column=0, sticky="e", padx=4)
        self.entry_user = tk.Entry(frame, width=22)
        self.entry_user.grid(row=0, column=1, padx=4, pady=2)

        tk.Label(frame, text="앱 비밀번호:").grid(row=0, column=2, sticky="e", padx=4)
        self.entry_pw = tk.Entry(frame, width=22)
        self.entry_pw.grid(row=0, column=3, padx=4, pady=2)

        self.var_remember = tk.BooleanVar(value=False)
        tk.Checkbutton(
            frame, text="앱 비밀번호 기억",
            variable=self.var_remember, command=self._on_remember_toggle,
        ).grid(row=0, column=4, padx=6)

        self.btn_login = tk.Button(frame, text="연결", width=9, command=self._on_login)
        self.btn_login.grid(row=0, column=5, padx=4)

        self.btn_logout = tk.Button(frame, text="연결 해제", width=9, command=self._on_logout)
        self.btn_logout.grid(row=0, column=6, padx=4)

        tk.Label(
            frame,
            text="※ 네이버 메일 설정 > POP3/IMAP 설정 > IMAP 사용함 으로 변경 후 앱 비밀번호를 사용하세요.",
            fg="gray", font=("", 8),
        ).grid(row=1, column=0, columnspan=7, sticky="w", pady=(2, 0))

    # ── 메일 검색 ─────────────────────────────────────────────────────────────

    def _build_search_frame(self):
        frame = tk.LabelFrame(self.root, text="메일 검색", padx=10, pady=6)
        frame.pack(fill="x", padx=10, pady=3)

        tk.Label(frame, text="보낸사람 아이디:").grid(row=0, column=0, sticky="e", padx=4)
        self.combo_sender = ttk.Combobox(frame, width=26)
        self.combo_sender.grid(row=0, column=1, padx=4, pady=2)
        self.combo_sender.bind("<Button-1>", lambda e: self._open_dropdown(self.combo_sender))

        tk.Label(frame, text="제목 키워드:").grid(row=0, column=2, sticky="e", padx=4)
        self.combo_subject = ttk.Combobox(frame, width=26)
        self.combo_subject.grid(row=0, column=3, padx=4, pady=2)
        self.combo_subject.bind("<Button-1>", lambda e: self._open_dropdown(self.combo_subject))

        self.btn_search = tk.Button(frame, text="검색", width=10, command=self._on_search)
        self.btn_search.grid(row=0, column=4, padx=8)

        tk.Label(frame, text="(빈칸이면 조건 무시)", fg="gray", font=("", 8)).grid(
            row=1, column=0, columnspan=5, sticky="w", padx=4
        )

    # ── 검색 결과 ─────────────────────────────────────────────────────────────

    def _build_result_frame(self):
        frame = tk.LabelFrame(self.root, text="검색 결과", padx=6, pady=4)
        frame.pack(fill="x", padx=10, pady=3)

        cols = ("선택", "날짜", "보낸사람", "제목", "첨부파일")
        self.tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="extended", height=10)
        col_widths = {"선택": 40, "날짜": 135, "보낸사람": 175, "제목": 340, "첨부파일": 100}
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=col_widths[c], anchor="center" if c == "선택" else "w")
        self.tree.bind("<Button-1>", self._on_tree_click)
        self.tree.bind("<Shift-Button-1>", self._on_tree_shift_click)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # 선택/다운로드 행 (결과 프레임 바깥)
        ctrl = tk.Frame(self.root)
        ctrl.pack(fill="x", padx=10, pady=(2, 3))

        self.btn_select_all = tk.Button(ctrl, text="전체 선택", width=9, command=self._select_all)
        self.btn_select_all.pack(side="left", padx=(0, 3))
        self.btn_deselect_all = tk.Button(ctrl, text="전체 해제", width=9, command=self._deselect_all)
        self.btn_deselect_all.pack(side="left")
        self.lbl_sel_count = tk.Label(ctrl, text="선택: 0개", fg="gray")
        self.lbl_sel_count.pack(side="left", padx=8)
        tk.Label(ctrl, text="(Shift+클릭 범위선택)", fg="gray", font=("", 8)).pack(side="left")

        self.btn_download = tk.Button(
            ctrl, text="첨부파일 다운로드", width=18, command=self._on_download
        )
        self.btn_download.pack(side="right", padx=4)
        tk.Label(ctrl, text=DESKTOP, fg="blue", font=("", 8, "underline")).pack(side="right")
        tk.Label(ctrl, text="저장:").pack(side="right")

    # ── 파일 변환 ─────────────────────────────────────────────────────────────

    def _build_convert_section(self):
        frame = tk.LabelFrame(self.root, text="파일 변환 (수탁업체명 교체)", padx=8, pady=6)
        frame.pack(fill="x", padx=10, pady=(0, 6))

        # 기준 파일 행
        ref_row = tk.Frame(frame)
        ref_row.pack(fill="x", pady=(0, 4))

        tk.Label(ref_row, text="기준 파일:").pack(side="left")
        self.ref_path_var = tk.StringVar(value="미설정 — 클릭하거나 드래그하세요")
        ref_lbl = tk.Label(
            ref_row, textvariable=self.ref_path_var,
            anchor="w", fg="#2563eb", relief="groove", padx=6, cursor="hand2",
        )
        ref_lbl.pack(side="left", fill="x", expand=True, padx=4)
        ref_lbl.bind("<Button-1>", lambda e: self._browse_ref_file())
        if HAS_DND:
            ref_lbl.drop_target_register(DND_FILES)
            ref_lbl.dnd_bind("<<Drop>>", self._on_ref_drop)

        tk.Button(ref_row, text="찾아보기", width=9,
                  command=self._browse_ref_file).pack(side="right", padx=(4, 0))
        self.ref_status_var = tk.StringVar(value="")
        tk.Label(ref_row, textvariable=self.ref_status_var,
                 fg="gray", font=("", 8), width=16).pack(side="right", padx=4)

        # 변환 파일 행 (리스트 + 버튼 세로)
        conv_row = tk.Frame(frame)
        conv_row.pack(fill="x")

        # 왼쪽: 파일 목록 (height=3) + 드래그 힌트
        left = tk.Frame(conv_row)
        left.pack(side="left", fill="both", expand=True)

        list_f = tk.Frame(left)
        list_f.pack(fill="x")
        self.conv_listbox = tk.Listbox(list_f, selectmode="extended", height=3)
        lsb = ttk.Scrollbar(list_f, orient="vertical", command=self.conv_listbox.yview)
        self.conv_listbox.configure(yscrollcommand=lsb.set)
        self.conv_listbox.pack(side="left", fill="both", expand=True)
        lsb.pack(side="right", fill="y")
        if HAS_DND:
            self.conv_listbox.drop_target_register(DND_FILES)
            self.conv_listbox.dnd_bind("<<Drop>>", self._on_conv_drop)

        hint_row = tk.Frame(left)
        hint_row.pack(fill="x", pady=(3, 0))
        hint = tk.Label(
            hint_row, text="변환할 파일을 여기에 드래그하세요 (.xlsx)",
            fg="#888", relief="groove", bd=1, padx=4, cursor="hand2", anchor="w",
        )
        hint.pack(side="left", fill="x", expand=True)
        hint.bind("<Button-1>", lambda e: self._browse_conv_files())
        if HAS_DND:
            hint.drop_target_register(DND_FILES)
            hint.dnd_bind("<<Drop>>", self._on_conv_drop)
        self.lbl_conv_count = tk.Label(hint_row, text="0개 파일", fg="gray", width=7)
        self.lbl_conv_count.pack(side="left", padx=4)

        # 오른쪽: 버튼 + 변환실행
        right = tk.Frame(conv_row)
        right.pack(side="right", padx=(8, 0), fill="y")

        tk.Button(right, text="파일 추가", width=9,
                  command=self._browse_conv_files).pack(pady=(0, 2))
        tk.Button(right, text="선택 삭제", width=9,
                  command=self._remove_conv_files).pack(pady=2)
        tk.Button(right, text="전체 삭제", width=9,
                  command=self._clear_conv_files).pack(pady=2)
        self.btn_convert = tk.Button(
            right, text="변환 실행", width=9,
            bg="#2563eb", fg="white", font=("", 9, "bold"),
            command=self._on_convert,
        )
        self.btn_convert.pack(pady=(6, 0))

    # ── 설정 ──────────────────────────────────────────────────────────────────

    def _load_config_into_ui(self):
        cfg = load_config()
        self._sender_history = cfg["sender_history"]
        self._subject_history = cfg["subject_history"]
        self.combo_sender["values"] = self._sender_history
        self.combo_subject["values"] = self._subject_history
        if self._sender_history:
            self.combo_sender.set(self._sender_history[0])
        if self._subject_history:
            self.combo_subject.set(self._subject_history[0])

        # 기준 파일 복원: 저장된 경로 → 없으면 실행 파일 옆 기본 파일
        ref_path = cfg["ref_file_path"]
        if not ref_path or not os.path.exists(ref_path):
            ref_path = get_default_ref_path()
        if os.path.exists(ref_path):
            self.ref_status_var.set("로딩 중...")
            threading.Thread(
                target=self._load_ref_file,
                args=(ref_path,),
                daemon=True,
            ).start()

        if cfg["remember"] and cfg["user"]:
            self.entry_user.insert(0, cfg["user"])
            self.entry_pw.insert(0, cfg["pw"])
            self.var_remember.set(True)
            self._remember = True
            self.root.after(200, self._on_login)

        self._apply_login_mode()

    def _persist_config(self):
        save_config(
            self._saved_user, self._saved_pw, self._remember,
            self._sender_history, self._subject_history,
            self._ref_file_path,
        )

    def _add_to_history(self, history, value):
        if not value:
            return history
        if value in history:
            history.remove(value)
        history.insert(0, value)
        return history[:HISTORY_MAX]

    def _on_remember_toggle(self):
        self._remember = self.var_remember.get()
        if not self._remember:
            clear_login_from_config()

    # ── 상태 관리 ──────────────────────────────────────────────────────────────

    def _apply_login_mode(self):
        for w in (self.entry_user, self.entry_pw):
            w.config(state="normal")
        self.btn_login.config(state="normal")
        self.btn_logout.config(state="disabled")
        self.btn_search.config(state="disabled")
        self.btn_download.config(state="disabled")
        self.btn_select_all.config(state="disabled")
        self.btn_deselect_all.config(state="disabled")

    def _apply_connected_mode(self):
        for w in (self.entry_user, self.entry_pw):
            w.config(state="disabled")
        self.btn_login.config(state="disabled")
        self.btn_logout.config(state="normal")
        self.btn_search.config(state="normal")
        self.btn_download.config(state="normal")
        self.btn_select_all.config(state="normal")
        self.btn_deselect_all.config(state="normal")

    def _set_busy(self, busy: bool):
        state = "disabled" if busy else "normal"
        self.btn_logout.config(state=state)
        self.btn_search.config(state=state)
        self.btn_download.config(state=state)

    def _status(self, msg):
        self.root.after(0, self.status_var.set, msg)

    def _build_status_bar(self):
        self.status_var = tk.StringVar(value="로그인해주세요.")
        tk.Label(
            self.root, textvariable=self.status_var,
            bd=1, relief="sunken", anchor="w", padx=6,
        ).pack(fill="x", side="bottom")

    # ── IMAP ──────────────────────────────────────────────────────────────────

    def _connect(self, user, pw):
        imap = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
        imap.sock.settimeout(SOCKET_TIMEOUT)
        login_id = f"{user}@naver.com" if "@" not in user else user
        imap.login(login_id, pw)
        return imap

    def _ensure_connected(self):
        with self._imap_lock:
            try:
                self.imap.noop()
            except Exception:
                self._status("연결이 끊겼습니다. 재연결 중...")
                try:
                    self.imap.logout()
                except Exception:
                    pass
                self.imap = self._connect(self._saved_user, self._saved_pw)
                self._status("재연결 성공.")

    def _select_inbox(self):
        with self._imap_lock:
            try:
                self.imap.select("INBOX")
            except Exception:
                self._status("연결이 끊겼습니다. 재연결 중...")
                try:
                    self.imap.logout()
                except Exception:
                    pass
                self.imap = self._connect(self._saved_user, self._saved_pw)
                self.imap.select("INBOX")
                self._status("재연결 성공.")

    # ── 로그인 ────────────────────────────────────────────────────────────────

    def _on_login(self):
        user = self.entry_user.get().strip()
        pw = self.entry_pw.get()
        if not user or not pw:
            messagebox.showwarning("입력 오류", "아이디와 앱 비밀번호를 입력하세요.")
            return
        self._status("연결 중...")
        self.btn_login.config(state="disabled")

        def connect():
            try:
                imap = self._connect(user, pw)
                with self._imap_lock:
                    self.imap = imap
                    self._saved_user = user   # 락 안에서 함께 갱신
                    self._saved_pw = pw
                self.root.after(0, self._login_success)
            except Exception as e:
                self.root.after(0, lambda: self._login_fail(str(e)))

        threading.Thread(target=connect, daemon=True).start()

    def _login_success(self):
        self._remember = self.var_remember.get()
        self._persist_config()
        self._apply_connected_mode()
        self._status("연결 성공. 검색 조건을 입력하고 검색하세요.")

    def _login_fail(self, err):
        self.btn_login.config(state="normal")
        messagebox.showerror(
            "연결 실패",
            f"로그인에 실패했습니다.\n\n{err}\n\n앱 비밀번호 또는 IMAP 활성화 여부를 확인하세요.",
        )
        self._status("연결 실패.")

    def _on_logout(self):
        with self._imap_lock:
            if self.imap:
                try:
                    self.imap.logout()
                except Exception:
                    pass
                self.imap = None
        self._saved_user = ""
        self._saved_pw = ""
        self._clear_results()
        self._apply_login_mode()
        self._status("연결이 해제되었습니다.")

    # ── 검색 ──────────────────────────────────────────────────────────────────

    def _open_dropdown(self, combo):
        if combo["values"]:
            combo.event_generate("<Down>")

    def _on_search(self):
        sender = self.combo_sender.get().strip()
        subject = self.combo_subject.get().strip()
        self._clear_results()
        self._status("검색 중...")
        self._set_busy(True)

        self._sender_history = self._add_to_history(self._sender_history, sender)
        self._subject_history = self._add_to_history(self._subject_history, subject)
        self.combo_sender["values"] = self._sender_history
        self.combo_subject["values"] = self._subject_history
        self._persist_config()

        def search():
            try:
                self._ensure_connected()
                results = self._fetch_mails(sender, subject)
                self.root.after(0, lambda: self._search_done(results))
            except Exception as e:
                self.root.after(0, lambda: self._search_fail(str(e)))

        threading.Thread(target=search, daemon=True).start()

    def _fetch_mails(self, sender, subject):
        self._select_inbox()
        criteria = []
        if sender and _is_ascii(sender):
            criteria.append(f'FROM "{sender}"')
        query = " ".join(criteria) if criteria else "ALL"

        with self._imap_lock:
            _, data = self.imap.search(None, query)
        uids = data[0].split()

        results = []
        for uid in reversed(uids[-500:]):
            try:
                with self._imap_lock:
                    _, msg_data = self.imap.fetch(uid, "(RFC822.HEADER RFC822.SIZE)")
                raw = msg_data[0][1]
                msg = email.message_from_bytes(raw)
                from_str = decode_header_value(msg.get("From", ""))
                subj_str = decode_header_value(msg.get("Subject", ""))
                if sender and not _is_ascii(sender) and sender.lower() not in from_str.lower():
                    continue
                if subject and subject.lower() not in subj_str.lower():
                    continue
                results.append({
                    "uid": uid,
                    "date": parse_date(msg.get("Date", "")),
                    "from": from_str,
                    "subject": subj_str,
                    "has_attach": "확인 필요",
                })
            except imaplib.IMAP4.abort:
                raise   # 연결 끊김은 상위로 전파
            except Exception:
                continue
        return results

    def _search_done(self, results):
        self.mail_list = results
        for item in results:
            self.tree.insert("", "end", values=(
                "☐", item["date"], item["from"], item["subject"], item["has_attach"]
            ))
        self.lbl_sel_count.config(text="선택: 0개")
        self._set_busy(False)
        self._status(f"검색 완료: {len(results)}개 메일")

    def _search_fail(self, err):
        self._set_busy(False)
        messagebox.showerror("검색 실패", err)
        self._status("검색 실패.")

    def _clear_results(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.mail_list = []
        self.lbl_sel_count.config(text="선택: 0개")

    # ── 트리뷰 선택 ────────────────────────────────────────────────────────────

    def _on_tree_click(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid:
            return "break"
        if iid in self.tree.selection():
            self.tree.selection_remove(iid)
        else:
            self.tree.selection_add(iid)
        self._update_checkmarks()
        return "break"

    def _on_tree_shift_click(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid:
            return "break"
        children = self.tree.get_children()
        selected = self.tree.selection()
        if selected:
            i1 = children.index(selected[-1])
            i2 = children.index(iid)
            lo, hi = min(i1, i2), max(i1, i2)
            for c in children[lo:hi + 1]:
                self.tree.selection_add(c)
        else:
            self.tree.selection_add(iid)
        self._update_checkmarks()
        return "break"

    def _update_checkmarks(self):
        selected = set(self.tree.selection())
        self.lbl_sel_count.config(text=f"선택: {len(selected)}개")
        for iid in self.tree.get_children():
            want = "☑" if iid in selected else "☐"
            vals = list(self.tree.item(iid, "values"))
            if vals[0] != want:
                vals[0] = want
                self.tree.item(iid, values=vals)

    def _select_all(self):
        self.tree.selection_set(self.tree.get_children())
        self._update_checkmarks()

    def _deselect_all(self):
        self.tree.selection_remove(self.tree.get_children())
        self._update_checkmarks()

    # ── 다운로드 ───────────────────────────────────────────────────────────────

    def _on_download(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("선택 없음", "다운로드할 메일을 선택하세요.")
            return
        mails = [self.mail_list[self.tree.index(s)] for s in selected]
        self._status("첨부파일 다운로드 중...")
        self._set_busy(True)

        def download():
            count, errors, saved_paths = 0, [], []
            for mail in mails:
                try:
                    n, paths = self._download_attachments(mail["uid"])
                    count += n
                    saved_paths.extend(paths)
                except Exception as e:
                    errors.append(str(e))
            self.root.after(0, lambda: self._download_done(count, errors, saved_paths))

        threading.Thread(target=download, daemon=True).start()

    def _download_attachments(self, uid):
        self._select_inbox()
        with self._imap_lock:
            _, msg_data = self.imap.fetch(uid, "(RFC822)")
        raw = msg_data[0][1]
        msg = email.message_from_bytes(raw)
        saved, paths = 0, []
        for part in msg.walk():
            filename = part.get_filename()
            if not filename:
                continue
            disposition = part.get("Content-Disposition", "")
            if "attachment" not in disposition and "inline" not in disposition:
                if part.get_content_maintype() == "multipart":
                    continue
            filename = self._safe_filename(decode_header_value(filename))
            payload = part.get_payload(decode=True)
            if payload is None:
                continue
            dest = os.path.join(DESKTOP, filename)
            base, ext = os.path.splitext(dest)
            counter = 1
            while os.path.exists(dest):
                dest = f"{base}_{counter}{ext}"
                counter += 1
            with open(dest, "wb") as f:
                f.write(payload)
            saved += 1
            paths.append(dest)
        return saved, paths

    @staticmethod
    def _safe_filename(name):
        for ch in r'\/:*?"<>|':
            name = name.replace(ch, "_")
        return name.strip()

    def _download_done(self, count, errors, saved_paths=None):
        self._set_busy(False)

        # 엑셀 파일만 변환 목록에 자동 추가
        excel_exts = {".xlsx", ".xlsm", ".xls"}
        added = 0
        for path in (saved_paths or []):
            if os.path.splitext(path)[1].lower() in excel_exts:
                self._add_conv_file(path)
                added += 1

        if errors:
            messagebox.showwarning("다운로드 부분 실패",
                                   f"{count}개 저장 완료.\n오류 {len(errors)}건\n{errors[0]}")
        elif count == 0:
            messagebox.showinfo("첨부파일 없음", "선택한 메일에 첨부파일이 없습니다.")
        else:
            msg = f"{count}개 첨부파일을 바탕화면에 저장했습니다."
            if added:
                msg += f"\n엑셀 파일 {added}개를 변환 목록에 추가했습니다."
            messagebox.showinfo("완료", msg)
        self._status(f"다운로드 완료: {count}개 저장 / 변환목록 {added}개 추가")

    # ── 파일 변환 탭 ───────────────────────────────────────────────────────────

    def _load_ref_file(self, path: str):
        """백그라운드 스레드에서 호출. 데이터 쓰기는 root.after로 메인 스레드에서 원자적 실행."""
        if not HAS_OPENPYXL:
            self.root.after(0, lambda: messagebox.showerror(
                "오류", "openpyxl 패키지가 필요합니다.\npip install openpyxl"))
            return
        try:
            ref_data = ReferenceData.load(path)

            def _apply(ref=ref_data, p=path):
                self._ref_data = ref
                self._ref_file_path = p
                self.ref_path_var.set(p)
                self.ref_status_var.set(f"✔ {len(ref.mapping)}개 업체 로드됨")
                self._persist_config()

            self.root.after(0, _apply)
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror(
                "기준 파일 오류", f"파일을 읽을 수 없습니다.\n{e}"))
            self.root.after(0, lambda: self.ref_status_var.set("✖ 파일 읽기 실패"))

    def _browse_ref_file(self):
        path = filedialog.askopenfilename(
            title="기준 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xlsm"), ("모든 파일", "*.*")],
        )
        if path:
            self._load_ref_file(path)

    def _on_ref_drop(self, event):
        paths = parse_dnd_paths(event.data)
        if paths:
            self._load_ref_file(paths[0])

    def _browse_conv_files(self):
        paths = filedialog.askopenfilenames(
            title="변환할 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xlsm"), ("모든 파일", "*.*")],
        )
        for p in paths:
            self._add_conv_file(p)

    def _on_conv_drop(self, event):
        for p in parse_dnd_paths(event.data):
            self._add_conv_file(p)

    def _add_conv_file(self, path: str):
        existing = list(self.conv_listbox.get(0, tk.END))
        if path not in existing:
            self.conv_listbox.insert(tk.END, path)
            self._update_conv_count()

    def _remove_conv_files(self):
        for idx in reversed(self.conv_listbox.curselection()):
            self.conv_listbox.delete(idx)
        self._update_conv_count()

    def _clear_conv_files(self):
        self.conv_listbox.delete(0, tk.END)
        self._update_conv_count()

    def _update_conv_count(self):
        n = self.conv_listbox.size()
        self.lbl_conv_count.config(text=f"{n}개 파일")

    def _on_convert(self):
        if not self._ref_data.mapping:
            messagebox.showwarning("기준 파일 없음", "기준 파일을 먼저 설정하세요.")
            return
        files = list(self.conv_listbox.get(0, tk.END))
        if not files:
            messagebox.showwarning("파일 없음", "변환할 파일을 추가하세요.")
            return

        self.btn_convert.config(state="disabled")
        self._status("파일 변환 중...")
        ref = self._ref_data  # 스냅샷 — 변환 중 기준 파일 재로드와 경쟁 없음

        def run():
            results = []
            for path in files:
                try:
                    out_path, changed, unmatched = convert_file(path, ref.mapping, ref.clinic, DESKTOP)
                    results.append((path, out_path, changed, unmatched, None))
                except Exception as e:
                    results.append((path, None, 0, [], str(e)))
            self.root.after(0, lambda: self._convert_done(results))

        threading.Thread(target=run, daemon=True).start()

    def _convert_done(self, results):
        self.btn_convert.config(state="normal")
        ok = [r for r in results if r[4] is None]
        fail = [r for r in results if r[4] is not None]

        lines = []
        total_unmatched = 0
        for path, out_path, changed, unmatched, err in ok:
            name = os.path.basename(out_path)
            lines.append(f"✔ {name}  ({changed}행 변환)")
            total_unmatched += len(unmatched)
        for path, _, _, _, err in fail:
            lines.append(f"✖ {os.path.basename(path)}: {err}")

        msg = "\n".join(lines) if lines else "변환 결과 없음"
        if total_unmatched:
            msg += f"\n\n매칭 실패 {total_unmatched}건 → 하단 목록에서 직접 입력하세요."
        if fail:
            messagebox.showwarning("변환 완료 (일부 실패)", msg)
        else:
            messagebox.showinfo("변환 완료", msg)
        self._status(f"변환 완료: 성공 {len(ok)}개, 실패 {len(fail)}개 / 미매칭 {total_unmatched}건")
        self._populate_unmatched(results)

    # ── 매칭 실패 패널 ─────────────────────────────────────────────────────────

    def _build_unmatched_section(self):
        self._unmatched_frame = tk.LabelFrame(
            self.root,
            text="매칭 실패 항목 — 수탁업체명 셀을 클릭해 입력 후 [수정 적용]",
            padx=8, pady=6,
        )
        # 변환 전까지는 숨김 (pack은 _populate_unmatched 에서 호출)

        tree_wrap = tk.Frame(self._unmatched_frame)
        tree_wrap.pack(fill="both", expand=True)

        cols = ("파일명", "행", "사업자번호 / 품목명", "현재 B열", "수탁업체명 입력")
        self.tree_unmatched = ttk.Treeview(
            tree_wrap, columns=cols, show="headings", height=6,
        )
        col_widths = {"파일명": 175, "행": 45, "사업자번호 / 품목명": 155, "현재 B열": 135, "수탁업체명 입력": 260}
        for c in cols:
            self.tree_unmatched.heading(c, text=c)
            self.tree_unmatched.column(
                c, width=col_widths[c], anchor="center" if c == "행" else "w"
            )
        self.tree_unmatched.tag_configure("clinic", background="#dbeafe")  # 연한 하늘색
        self.tree_unmatched.bind("<Button-1>", self._on_unmatched_click)

        vsb_u = ttk.Scrollbar(tree_wrap, orient="vertical", command=self.tree_unmatched.yview)
        self.tree_unmatched.configure(yscrollcommand=vsb_u.set)
        vsb_u.pack(side="right", fill="y")
        self.tree_unmatched.pack(side="left", fill="both", expand=True)

        hint = tk.Label(
            self._unmatched_frame,
            text="※ 마지막 열(수탁업체명 입력)을 클릭하면 직접 입력할 수 있습니다.",
            fg="gray", font=("", 8),
        )
        hint.pack(anchor="w", pady=(2, 0))

        btn_row = tk.Frame(self._unmatched_frame)
        btn_row.pack(fill="x", pady=(4, 0))

        tk.Button(
            btn_row, text="수정 적용", width=10,
            bg="#16a34a", fg="white", font=("", 9, "bold"),
            command=self._apply_corrections,
        ).pack(side="left", padx=(0, 4))
        tk.Button(
            btn_row, text="목록 지우기", width=10,
            command=self._clear_unmatched,
        ).pack(side="left")
        self.lbl_unmatched_count = tk.Label(btn_row, text="", fg="gray")
        self.lbl_unmatched_count.pack(side="right")

    def _fit_window(self):
        """콘텐츠에 맞게 창 높이를 자동 조정."""
        self.root.update_idletasks()
        w = self.root.winfo_width()
        self.root.geometry(f"{w}x{self.root.winfo_reqheight()}")

    def _populate_unmatched(self, results):
        for iid in self.tree_unmatched.get_children():
            self.tree_unmatched.delete(iid)
        self._unmatched_data.clear()

        rows_to_add = []
        for path, out_path, changed, unmatched, err in results:
            if err or not unmatched:
                continue
            label = os.path.basename(out_path)
            for item in unmatched:
                rows_to_add.append((label, out_path, item))

        if not rows_to_add:
            self._unmatched_frame.pack_forget()
            self._fit_window()
            return

        for label, out_path, item in rows_to_add:
            is_clinic = item["biz_raw"].startswith("[윤병옥]")
            tag = ("clinic",) if is_clinic else ()
            iid = self.tree_unmatched.insert("", "end", values=(
                label, item["row"], item["biz_raw"], item["current_b"], "",
            ), tags=tag)
            self._unmatched_data[iid] = {"out_path": out_path, "row": item["row"], "biz_raw": item["biz_raw"]}

        total = len(rows_to_add)
        self.lbl_unmatched_count.config(text=f"총 {total}건 / 미입력 {total}건")
        self._unmatched_frame.pack(fill="x", padx=10, pady=(0, 6))
        self._fit_window()

    def _on_unmatched_click(self, event):
        region = self.tree_unmatched.identify_region(event.x, event.y)
        col = self.tree_unmatched.identify_column(event.x)
        iid = self.tree_unmatched.identify_row(event.y)
        if region == "cell" and col == "#5" and iid:
            self._start_cell_edit(iid)

    def _start_cell_edit(self, iid):
        if self._edit_entry is not None:
            try:
                self._edit_entry.destroy()
            except Exception:
                pass
            self._edit_entry = None

        bbox = self.tree_unmatched.bbox(iid, "#5")
        if not bbox:
            return
        x, y, w, h = bbox

        current = self.tree_unmatched.set(iid, "수탁업체명 입력")
        entry = ttk.Combobox(
            self.tree_unmatched,
            values=self._ref_data.names,
            font=("", 9),
        )
        entry.place(x=x, y=y, width=w, height=h)
        entry.set(current)
        entry.focus_set()
        entry.event_generate("<Down>")
        self._edit_entry = entry

        _done = [False]

        def commit(event=None):
            if _done[0]:
                return
            _done[0] = True
            val = entry.get().strip()
            self.tree_unmatched.set(iid, "수탁업체명 입력", val)
            try:
                entry.destroy()
            except Exception:
                pass
            self._edit_entry = None
            self._update_unmatched_count()

        def cancel(event=None):
            if _done[0]:
                return
            _done[0] = True
            try:
                entry.destroy()
            except Exception:
                pass
            self._edit_entry = None

        def on_focus_out(event=None):
            # ComboboxSelected 가 먼저 처리되도록 150ms 대기
            entry.after(150, commit)

        def on_key_release(event=None):
            if event and event.keysym in (
                "Return", "Tab", "Escape", "BackSpace", "Delete",
                "Left", "Right", "Up", "Down", "Home", "End",
            ):
                return
            typed = entry.get()
            if not typed:
                entry["values"] = self._ref_data.names
                return
            low = typed.lower()
            matches = [n for n in self._ref_data.names if low in n.lower()]
            entry["values"] = matches if matches else self._ref_data.names
            # entry.set() 은 한글 IME 조합 중 버퍼 리셋으로 글자 중복 유발 — 제거
            # 유일 매칭이면 드롭다운만 열어 사용자가 선택하도록 유도
            if len(matches) == 1:
                entry.event_generate("<Down>")

        entry.bind("<<ComboboxSelected>>", commit)
        entry.bind("<Return>", commit)
        entry.bind("<Tab>", commit)
        entry.bind("<FocusOut>", on_focus_out)
        entry.bind("<Escape>", cancel)
        entry.bind("<KeyRelease>", on_key_release)

    def _update_unmatched_count(self):
        children = self.tree_unmatched.get_children()
        total = len(children)
        empty = sum(
            1 for iid in children
            if not self.tree_unmatched.set(iid, "수탁업체명 입력").strip()
        )
        self.lbl_unmatched_count.config(text=f"총 {total}건 / 미입력 {empty}건")

    def _clear_unmatched(self):
        for iid in self.tree_unmatched.get_children():
            self.tree_unmatched.delete(iid)
        self._unmatched_data.clear()
        self._unmatched_frame.pack_forget()
        self._fit_window()

    def _apply_corrections(self):
        corrections: dict = {}
        ref_additions: list = []
        applied_iids = []
        for iid in self.tree_unmatched.get_children():
            new_val = self.tree_unmatched.set(iid, "수탁업체명 입력").strip()
            if not new_val:
                continue
            meta = self._unmatched_data.get(iid)
            if not meta:
                continue
            corrections.setdefault(meta["out_path"], []).append((meta["row"], new_val))
            ref_additions.append((meta.get("biz_raw", ""), new_val))
            applied_iids.append(iid)

        if not corrections:
            messagebox.showwarning(
                "입력 없음",
                "수정할 수탁업체명을 입력하세요.\n마지막 열 셀을 클릭하면 입력할 수 있습니다.",
            )
            return

        self.btn_convert.config(state="disabled")
        self._status("수정 적용 중...")
        ref_path = self._ref_file_path

        def run():
            errors = []
            succeeded_paths: set = set()
            for out_path, rows in corrections.items():
                try:
                    self._write_corrections(out_path, rows)
                    succeeded_paths.add(out_path)
                except Exception as e:
                    errors.append(f"{os.path.basename(out_path)}: {e}")

            ref_ok = False
            if ref_additions and ref_path and os.path.exists(ref_path):
                try:
                    _append_to_reference(ref_path, ref_additions)
                    ref_ok = True
                except Exception as e:
                    errors.append(f"기준파일 업데이트 실패: {e}")

            total = sum(len(v) for k, v in corrections.items() if k in succeeded_paths)
            self.root.after(
                0,
                lambda: self._corrections_done(errors, total, applied_iids, succeeded_paths, ref_ok),
            )

        threading.Thread(target=run, daemon=True).start()

    def _corrections_done(self, errors, total, applied_iids, succeeded_paths, ref_ok=False):
        self.btn_convert.config(state="normal")
        if errors:
            messagebox.showwarning(
                "수정 완료 (일부 실패)",
                f"{total}건 적용 완료.\n\n오류:\n" + "\n".join(errors),
            )
        else:
            suffix = "\n기준파일에도 추가되었습니다." if ref_ok else ""
            messagebox.showinfo("수정 완료", f"{total}건 수정 완료.{suffix}")

        if ref_ok and self._ref_file_path:
            threading.Thread(
                target=self._load_ref_file,
                args=(self._ref_file_path,),
                daemon=True,
            ).start()

        for iid in applied_iids:
            meta = self._unmatched_data.get(iid)
            if meta and meta["out_path"] in succeeded_paths:
                try:
                    self.tree_unmatched.delete(iid)
                    self._unmatched_data.pop(iid, None)
                except Exception:
                    pass

        remaining = self.tree_unmatched.get_children()
        if not remaining:
            self._unmatched_frame.pack_forget()
            self._fit_window()
        else:
            self._update_unmatched_count()
        self._status(f"수정 적용 완료: {total}건")

    def _write_corrections(self, out_path: str, rows: list) -> int:
        return get_excel_backend().write_cells(out_path, rows)


def main():
    try:
        import ctypes
        ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)
    except Exception:
        pass

    if HAS_DND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    NaverMailApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
