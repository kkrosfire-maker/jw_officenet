import openpyxl

FIELDS = ["성분", "보험코드", "제약사명", "제품명", "약가", "요율", "비고"]


def _norm(s):
    if s is None:
        return ""
    return str(s).replace("\n", "").replace(" ", "").strip()


def _find_header(ws, max_scan_rows=15):
    """첫 max_scan_rows행 안에서 필요한 필드명이 가장 많이 일치하는 행을 헤더로 판단한다."""
    best = None
    best_score = 0
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
    ):
        mapping = {}
        for c_idx, val in enumerate(row, start=1):
            key = _norm(val)
            for field in FIELDS:
                if key == field:
                    mapping[field] = c_idx
        score = len(mapping)
        if score > best_score:
            best_score = score
            best = (r_idx, mapping)
    if best and "제품명" in best[1] and "성분" in best[1]:
        return best
    return None


def _get_font_color(cell):
    """셀의 글자색을 '#RRGGBB' 문자열로 반환한다. 기본(검정) 색이면 None."""
    font = cell.font
    if font is None or font.color is None:
        return None
    color = font.color
    if color.type != "rgb":
        return None
    rgb = color.rgb
    if not isinstance(rgb, str) or len(rgb) < 6:
        return None
    rgb = rgb[-6:].upper()
    if rgb == "000000":
        return None
    return "#" + rgb


def load_rate_sheet(path):
    """
    기준 엑셀 파일을 읽어 (rows, field_map, sheet_name)을 반환한다.
    rows: 각 항목이 {field: value} 인 dict 리스트 (FIELDS 중 실제 존재하는 필드만 포함)
    각 row에는 원본의 글자색을 담은 "_row_color"("#RRGGBB" 또는 None)도 포함된다.
    field_map: {field: 원본 열 번호} - 실제 발견된 필드만 포함
    시트가 여러 개면 필요한 헤더가 가장 잘 매칭되는 시트를 자동 선택한다.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        best_sheet = None
        best_header = None
        best_score = 0
        for name in wb.sheetnames:
            ws = wb[name]
            found = _find_header(ws)
            if found and len(found[1]) > best_score:
                best_score = len(found[1])
                best_sheet = name
                best_header = found

        if best_header is None:
            raise ValueError(
                "필요한 컬럼(제품명, 성분 등)을 찾을 수 없습니다.\n"
                "엑셀 헤더에 '제품명'과 '성분' 열이 있는지 확인해주세요."
            )

        header_row, field_map = best_header
        col_to_field = {c: f for f, c in field_map.items()}
        ws = wb[best_sheet]

        rows = []
        for row_cells in ws.iter_rows(min_row=header_row + 1):
            row_vals = {}
            has_data = False
            row_color = None
            for c_idx, field in col_to_field.items():
                cell = row_cells[c_idx - 1] if c_idx - 1 < len(row_cells) else None
                v = cell.value if cell is not None else None
                if v is not None and v != "":
                    has_data = True
                row_vals[field] = v
                if row_color is None and cell is not None:
                    row_color = _get_font_color(cell)
            if has_data:
                row_vals["_row_color"] = row_color
                rows.append(row_vals)

        return rows, field_map, best_sheet
    finally:
        wb.close()


def search_rows(rows, keyword, search_fields, exact=False):
    """
    search_fields(제품명/성분 등) 중 하나라도 keyword를 포함(or 정확히 일치)하면 결과에 포함.
    keyword는 콤마(,)로 여러 단어를 구분해 입력할 수 있으며, 그 중 하나라도 매칭되면 포함(OR 검색).
    """
    raw = (keyword or "").strip()
    if not raw:
        return list(rows)

    terms = [t.strip() for t in raw.split(",") if t.strip()]
    if not terms:
        return list(rows)
    terms_lower = [t.lower() for t in terms]

    results = []
    for row in rows:
        matched = False
        for field in search_fields:
            v = row.get(field)
            if v is None:
                continue
            s = str(v)
            s_lower = s.lower()
            for term, term_lower in zip(terms, terms_lower):
                if exact:
                    if s.strip() == term:
                        matched = True
                        break
                else:
                    if term_lower in s_lower:
                        matched = True
                        break
            if matched:
                break
        if matched:
            results.append(row)
    return results


def format_value(field, value):
    """표시용 문자열 포맷팅."""
    if value is None:
        return ""
    if field == "요율":
        if isinstance(value, (int, float)):
            return f"{value * 100:.1f}%"
        return str(value)
    if field == "약가":
        if isinstance(value, (int, float)):
            if float(value).is_integer():
                return f"{int(value):,}"
            return f"{value:,.2f}"
        return str(value)
    return str(value)
