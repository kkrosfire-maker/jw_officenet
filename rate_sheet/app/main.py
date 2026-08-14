import ctypes
import os
import re
import threading
import time
import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, filedialog, messagebox

from config import load_config, save_config
from excel_utils import FIELDS, load_rate_sheet, search_rows, format_value

SEARCH_TARGET_FIELDS = ["제품명", "성분"]


class _LOGFONTW(ctypes.Structure):
    _fields_ = [
        ("lfHeight", ctypes.c_long),
        ("lfWidth", ctypes.c_long),
        ("lfEscapement", ctypes.c_long),
        ("lfOrientation", ctypes.c_long),
        ("lfWeight", ctypes.c_long),
        ("lfItalic", ctypes.c_byte),
        ("lfUnderline", ctypes.c_byte),
        ("lfStrikeOut", ctypes.c_byte),
        ("lfCharSet", ctypes.c_byte),
        ("lfOutPrecision", ctypes.c_byte),
        ("lfClipPrecision", ctypes.c_byte),
        ("lfQuality", ctypes.c_byte),
        ("lfPitchAndFamily", ctypes.c_byte),
        ("lfFaceName", ctypes.c_wchar * 32),
    ]


def sync_ime_composition_font(widget, family, point_size):
    """
    Windows는 위젯에 커밋된(완성된) 글자는 위젯 자체 폰트로 그리지만,
    한글 IME로 조합 중인(preedit) 글자는 별도의 "조합 폰트"로 그린다.
    Tk가 이 조합 폰트를 위젯 폰트와 동기화해주지 않아서, 타이핑 중엔 IME의
    기본(더 큰) 폰트로 보이다가 조합이 끝나 커밋되는 순간 위젯 폰트 크기로
    돌아오는 현상이 생긴다. 포커스를 받을 때마다 조합 폰트를 위젯 폰트와
    맞춰서 이 크기 불일치를 없앤다.
    """
    try:
        hwnd = widget.winfo_id()
        imm32 = ctypes.WinDLL("imm32")
        user32 = ctypes.WinDLL("user32")
        gdi32 = ctypes.WinDLL("gdi32")

        hdc = user32.GetDC(hwnd)
        LOGPIXELSY = 90
        dpi = gdi32.GetDeviceCaps(hdc, LOGPIXELSY) or 96
        user32.ReleaseDC(hwnd, hdc)

        logfont = _LOGFONTW()
        logfont.lfHeight = -int(round(point_size * dpi / 72))
        logfont.lfWeight = 400
        logfont.lfCharSet = 129  # HANGUL_CHARSET
        logfont.lfFaceName = family

        himc = imm32.ImmGetContext(hwnd)
        if himc:
            imm32.ImmSetCompositionFontW(himc, ctypes.byref(logfont))
            imm32.ImmReleaseContext(hwnd, himc)
    except Exception:
        pass
CHECK_COL = "선택"
CHECK_MARK = "☑"
UNCHECK_MARK = "☐"
HEADER_BG = "#15803D"
HEADER_FG = "#FFFFFF"
SECTION_ORDER = ["file", "search", "target", "columns", "result", "bottom"]
SECTION_TITLES = {
    "file": "기준 엑셀",
    "search": "검색어",
    "target": "검색 대상",
    "columns": "표시할 항목",
    "result": "결과",
    "bottom": "도구",
}

# ---- 디자인 토큰 ----
APP_BG = "#F1F2F4"       # 창 배경(캔버스)
CARD_BG = "#FFFFFF"      # 섹션(카드) 배경
BORDER = "#E1E4E8"       # 카드 테두리
TEXT = "#1F2328"         # 기본 텍스트
MUTED_TEXT = "#6B7280"   # 보조/힌트 텍스트
ACCENT = "#15803D"       # 브랜드 포인트(초록)
ACCENT_HOVER = "#116C31"
HANDLE_BG = "#F6F7F8"    # 섹션 손잡이 바 배경
HANDLE_FG = "#57606A"    # 섹션 손잡이 바 텍스트
BASE_FONT = ("맑은 고딕", 9)


class RateSheetApp:
    def __init__(self, root):
        self.root = root
        self.root.title("요율표 검색기")
        self.root.geometry("1100x780")

        self.cfg = load_config()
        self.rows = []
        self.field_map = {}
        self.current_results = []
        self.current_columns = []
        self.sort_col = None
        self.sort_ascending = True
        self._known_color_tags = set()
        self._sort_after_id = None
        self._grid_lines = []
        self._grid_redraw_after_id = None
        self._header_height = 28
        self._edit_entry = None
        self._edit_row_id = None

        self.file_path_var = tk.StringVar(value=self.cfg.get("last_file", ""))
        self.keyword_var = tk.StringVar()
        self.exact_var = tk.BooleanVar(value=False)
        self.status_var = tk.StringVar(value="기준 엑셀 파일을 선택해주세요.")

        self.search_target_vars = {
            f: tk.BooleanVar(value=True) for f in SEARCH_TARGET_FIELDS
        }
        self.col_vars = {f: tk.BooleanVar(value=True) for f in FIELDS}
        self.select_all_var = tk.BooleanVar(value=True)

        self._build_widgets()

        if self.file_path_var.get() and os.path.exists(self.file_path_var.get()):
            self._load_file(self.file_path_var.get())

    # ---------- UI 구성 ----------
    def _build_widgets(self):
        self.root.configure(background=APP_BG)

        style = ttk.Style()
        style.theme_use("clam")

        style.configure(".", background=CARD_BG, foreground=TEXT, font=BASE_FONT)
        style.configure("TFrame", background=CARD_BG)
        style.configure("TLabel", background=CARD_BG, foreground=TEXT)
        style.configure("TCheckbutton", background=CARD_BG, foreground=TEXT)
        style.map(
            "TCheckbutton",
            background=[("active", CARD_BG)],
            foreground=[("active", TEXT)],
        )
        style.configure("TEntry", font=BASE_FONT)

        style.configure(
            "Card.TFrame", background=CARD_BG, relief="solid", borderwidth=1
        )
        style.configure(
            "Handle.TFrame", background=HANDLE_BG, relief="flat", borderwidth=0
        )

        style.configure(
            "TButton",
            padding=(10, 5),
            font=BASE_FONT,
            background="#F0F1F3",
            foreground=TEXT,
        )
        style.map("TButton", background=[("active", "#E4E6E9")])

        style.configure(
            "Accent.TButton",
            padding=(12, 5),
            font=(BASE_FONT[0], BASE_FONT[1], "bold"),
            background=ACCENT,
            foreground="#FFFFFF",
        )
        style.map("Accent.TButton", background=[("active", ACCENT_HOVER)])

        style.configure(
            "Treeview.Heading",
            background=HEADER_BG,
            foreground=HEADER_FG,
            font=(BASE_FONT[0], BASE_FONT[1], "bold"),
            relief="flat",
        )
        style.map("Treeview.Heading", background=[("active", HEADER_BG)])
        style.configure(
            "Treeview",
            rowheight=26,
            borderwidth=1,
            relief="solid",
            bordercolor="#B0B0B0",
            font=BASE_FONT,
        )

        self.section_order = list(SECTION_ORDER)
        self.default_section_order = list(SECTION_ORDER)
        self.section_frames = {}
        self._drag_name = None

        self._build_file_section()
        self._build_search_section()
        self._build_target_section()
        self._build_columns_section()
        self._build_result_section()
        self._build_bottom_section()

        self._repack_sections()
        self._apply_min_size()

    def _apply_min_size(self):
        """도구 섹션 등 고정 섹션이 창 높이 부족으로 잘리지 않도록 최소 크기를 강제한다."""
        self.root.update_idletasks()
        fixed_height = 0
        for name in self.section_order:
            if name == "result":
                continue
            fixed_height += self.section_frames[name].winfo_reqheight()

        min_height = fixed_height + 220
        min_width = 900
        self.root.minsize(min_width, min_height)

        if self.root.winfo_height() < min_height:
            current_width = max(self.root.winfo_width(), min_width)
            self.root.geometry(f"{current_width}x{min_height}")

    def _make_section(self, name):
        outer = ttk.Frame(self.root, style="Card.TFrame")
        handle = tk.Label(
            outer,
            text=f"⠿ {SECTION_TITLES[name]}",
            anchor="w",
            background=HANDLE_BG,
            foreground=HANDLE_FG,
            font=BASE_FONT,
            cursor="fleur",
            padx=8,
            pady=3,
        )
        handle.pack(fill="x")
        handle.bind("<ButtonPress-1>", lambda e, n=name: self._on_drag_start(e, n))
        handle.bind("<ButtonRelease-1>", self._on_drag_release)

        content = ttk.Frame(outer, style="TFrame")
        content.pack(fill="both", expand=True, padx=8, pady=6)

        self.section_frames[name] = outer
        return content

    def _repack_sections(self):
        for name in self.section_order:
            self.section_frames[name].pack_forget()
        for name in self.section_order:
            frame = self.section_frames[name]
            if name == "result":
                frame.pack(fill="both", expand=True, padx=6, pady=3)
            else:
                frame.pack(fill="x", padx=6, pady=3)

    def _reset_section_order(self):
        if self.section_order != self.default_section_order:
            self.section_order = list(self.default_section_order)
            self._repack_sections()

    def _on_drag_start(self, event, name):
        self._drag_name = name

    def _on_drag_release(self, event):
        name = self._drag_name
        self._drag_name = None
        if not name:
            return

        drop_y = event.y_root
        centers = []
        for n in self.section_order:
            frame = self.section_frames[n]
            top = frame.winfo_rooty()
            height = frame.winfo_height()
            centers.append((n, top + height / 2))

        target_index = len(centers)
        for i, (_, center) in enumerate(centers):
            if drop_y < center:
                target_index = i
                break

        current_index = self.section_order.index(name)
        if target_index == current_index or target_index == current_index + 1:
            return

        self.section_order.pop(current_index)
        if target_index > current_index:
            target_index -= 1
        self.section_order.insert(target_index, name)
        self._repack_sections()

    def _build_file_section(self):
        content = self._make_section("file")
        ttk.Label(content, text="기준 엑셀:").pack(side="left")
        ttk.Entry(
            content, textvariable=self.file_path_var, state="readonly"
        ).pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(content, text="파일 선택", command=self._browse_file).pack(
            side="left"
        )

    def _build_search_section(self):
        content = self._make_section("search")
        ttk.Label(content, text="검색어:").pack(side="left")
        entry = ttk.Entry(
            content, textvariable=self.keyword_var, font=("맑은 고딕", 10)
        )
        entry.pack(side="left", fill="x", expand=True, padx=6)
        entry.bind("<Return>", lambda e: self._do_search())
        entry.bind(
            "<FocusIn>",
            lambda e: sync_ime_composition_font(entry, "맑은 고딕", 10),
        )
        ttk.Button(
            content, text="검색", command=self._do_search, style="Accent.TButton"
        ).pack(side="left")
        ttk.Label(
            content, text="(쉼표로 구분해 여러 단어 검색 가능: 예) 타이레놀, 아스피린)",
            foreground=MUTED_TEXT,
        ).pack(side="left", padx=(10, 0))

    def _build_target_section(self):
        content = self._make_section("target")
        ttk.Label(content, text="검색 대상:").pack(side="left")
        for f in SEARCH_TARGET_FIELDS:
            ttk.Checkbutton(
                content, text=f, variable=self.search_target_vars[f]
            ).pack(side="left", padx=4)
        ttk.Checkbutton(
            content, text="정확히 일치", variable=self.exact_var
        ).pack(side="left", padx=(20, 4))

    def _build_columns_section(self):
        content = self._make_section("columns")
        ttk.Checkbutton(
            content,
            text="전체선택",
            variable=self.select_all_var,
            command=self._toggle_select_all,
        ).pack(side="left", padx=(4, 16))

        self.columns_content = content
        self.column_order = list(FIELDS)
        self.column_item_frames = {}
        self._col_drag_name = None
        for f in self.column_order:
            self._make_column_item(f)

    def _make_column_item(self, f):
        item = ttk.Frame(self.columns_content, style="TFrame")
        handle = tk.Label(
            item,
            text="⠿",
            background=CARD_BG,
            foreground=MUTED_TEXT,
            cursor="fleur",
            padx=2,
        )
        handle.pack(side="left")
        handle.bind("<ButtonPress-1>", lambda e, n=f: self._on_col_drag_start(e, n))
        handle.bind("<ButtonRelease-1>", self._on_col_drag_release)
        ttk.Checkbutton(item, text=f, variable=self.col_vars[f]).pack(side="left")
        item.pack(side="left", padx=4)
        self.column_item_frames[f] = item

    def _repack_column_items(self):
        for f in self.column_order:
            self.column_item_frames[f].pack_forget()
        for f in self.column_order:
            self.column_item_frames[f].pack(side="left", padx=4)

    def _on_col_drag_start(self, event, name):
        self._col_drag_name = name

    def _on_col_drag_release(self, event):
        name = self._col_drag_name
        self._col_drag_name = None
        if not name:
            return

        drop_x = event.x_root
        centers = []
        for n in self.column_order:
            frame = self.column_item_frames[n]
            left = frame.winfo_rootx()
            width = frame.winfo_width()
            centers.append((n, left + width / 2))

        target_index = len(centers)
        for i, (_, center) in enumerate(centers):
            if drop_x < center:
                target_index = i
                break

        current_index = self.column_order.index(name)
        if target_index == current_index or target_index == current_index + 1:
            return

        self.column_order.pop(current_index)
        if target_index > current_index:
            target_index -= 1
        self.column_order.insert(target_index, name)
        self._repack_column_items()

    def _build_result_section(self):
        content = self._make_section("result")
        self.result_content = content

        self.tree = ttk.Treeview(content, show="headings")
        self.tree.tag_configure("evenrow", background="#FFFFFF")
        self.tree.tag_configure("oddrow", background="#F5F6F7")
        self.tree.bind("<ButtonRelease-1>", self._on_heading_click)
        self.tree.bind("<Double-Button-1>", self._on_tree_double_click)
        self.tree.bind("<Configure>", self._schedule_grid_redraw)
        vsb = ttk.Scrollbar(content, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(content, orient="horizontal", command=self.tree.xview)
        self._hsb = hsb
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=self._on_xscroll_set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        content.rowconfigure(0, weight=1)
        content.columnconfigure(0, weight=1)

    def _build_bottom_section(self):
        content = self._make_section("bottom")
        ttk.Button(
            content, text="엑셀로 내보내기", command=self._export_excel
        ).pack(side="left")
        ttk.Button(
            content, text="클립보드 복사", command=self._copy_clipboard
        ).pack(side="left", padx=6)
        ttk.Button(
            content, text="캡쳐 후 복사", command=self._capture_copy
        ).pack(side="left", padx=6)
        ttk.Label(content, textvariable=self.status_var).pack(side="left", padx=16)

    def _toggle_select_all(self):
        val = self.select_all_var.get()
        for var in self.col_vars.values():
            var.set(val)

    # ---------- 파일 로드 ----------
    def _browse_file(self):
        initialdir = (
            os.path.dirname(self.file_path_var.get())
            if self.file_path_var.get()
            else None
        )
        path = filedialog.askopenfilename(
            title="기준 엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xlsm *.xls")],
            initialdir=initialdir,
        )
        if path:
            self._load_file(path)

    def _load_file(self, path):
        self.status_var.set("파일 불러오는 중...")
        self.root.config(cursor="watch")
        self.root.update_idletasks()

        def worker():
            try:
                rows, field_map, sheet_name = load_rate_sheet(path)
                self.root.after(
                    0, lambda: self._on_file_loaded(path, rows, field_map, sheet_name)
                )
            except Exception as e:
                self.root.after(0, lambda: self._on_file_load_error(e))

        threading.Thread(target=worker, daemon=True).start()

    def _on_file_loaded(self, path, rows, field_map, sheet_name):
        self.rows = rows
        self.field_map = field_map
        self.file_path_var.set(path)
        self.cfg["last_file"] = path
        save_config(self.cfg)
        self.root.config(cursor="")
        missing = [f for f in FIELDS if f not in field_map]
        msg = f"'{sheet_name}' 시트, {len(rows)}건 로드 완료"
        if missing:
            msg += f" (누락 항목: {', '.join(missing)})"
        self.status_var.set(msg)

    def _on_file_load_error(self, error):
        self.root.config(cursor="")
        self.status_var.set("파일 로드 실패")
        messagebox.showerror("오류", str(error))

    # ---------- 검색 ----------
    def _do_search(self):
        if not self.rows:
            messagebox.showwarning("알림", "먼저 기준 엑셀 파일을 선택해주세요.")
            return

        search_fields = [f for f, v in self.search_target_vars.items() if v.get()]
        if not search_fields:
            messagebox.showwarning("알림", "검색 대상을 하나 이상 선택해주세요.")
            return

        columns = [f for f in self.column_order if self.col_vars[f].get()]
        if not columns:
            messagebox.showwarning("알림", "표시할 항목을 하나 이상 선택해주세요.")
            return

        self._reset_section_order()

        keyword = self.keyword_var.get()
        results = search_rows(self.rows, keyword, search_fields, self.exact_var.get())

        self.current_results = results
        self.current_columns = columns
        self.sort_col = None
        self.sort_ascending = True
        self._setup_columns(columns)
        self._refresh_rows(results, columns)
        self.status_var.set(f"검색 결과: {len(results)}건")

    def _setup_columns(self, columns):
        self.tree["columns"] = [CHECK_COL] + columns
        self.tree.column(CHECK_COL, width=36, anchor="center", stretch=False)
        for col in columns:
            self.tree.heading(col, text=col)
            width = 260 if col in ("제품명", "성분", "비고") else 110
            anchor = "center" if col in ("보험코드", "약가", "요율") else "w"
            self.tree.column(col, width=width, anchor=anchor, stretch=False)

    def _refresh_rows(self, results, columns):
        self._cancel_edit()
        self.tree.delete(*self.tree.get_children())
        for idx, row in enumerate(results):
            mark = CHECK_MARK if row.get("_checked", True) else UNCHECK_MARK
            values = [mark] + [format_value(f, row.get(f)) for f in columns]
            tags = ["evenrow" if idx % 2 == 0 else "oddrow"]
            color = row.get("_row_color")
            if color:
                tags.append(self._color_tag(color))
            self.tree.insert("", "end", values=values, tags=tuple(tags))
        self._update_heading_arrows()
        self._update_check_header()
        self._redraw_grid_lines()

    def _update_check_header(self):
        if self.current_results and all(
            r.get("_checked", True) for r in self.current_results
        ):
            mark = CHECK_MARK
        else:
            mark = UNCHECK_MARK
        self.tree.heading(CHECK_COL, text=mark)

    def _color_tag(self, hex_color):
        tag = f"color_{hex_color.lstrip('#')}"
        if tag not in self._known_color_tags:
            self.tree.tag_configure(tag, foreground=hex_color)
            self._known_color_tags.add(tag)
        return tag

    def _update_heading_arrows(self):
        for col in self.current_columns:
            arrow = ""
            if col == self.sort_col:
                arrow = " ▲" if self.sort_ascending else " ▼"
            self.tree.heading(col, text=col + arrow)

    def _sort_by(self, col):
        if not self.current_results:
            return
        if self.sort_col == col:
            self.sort_ascending = not self.sort_ascending
        else:
            self.sort_col = col
            self.sort_ascending = True

        def sort_key(row):
            v = row.get(col)
            if col in ("약가", "요율"):
                return float(v) if isinstance(v, (int, float)) else 0.0
            return str(v) if v is not None else ""

        def is_missing(row):
            v = row.get(col)
            if col in ("약가", "요율"):
                return not isinstance(v, (int, float))
            return v is None or v == ""

        # 값 기준으로 우선 정렬한 뒤, 값이 없는(N/A) 행은 방향과 무관하게 항상 맨 아래로 보낸다.
        self.current_results.sort(key=sort_key, reverse=not self.sort_ascending)
        self.current_results.sort(key=is_missing)
        self._refresh_rows(self.current_results, self.current_columns)

    def _on_heading_click(self, event):
        """
        헤더 단일 클릭 -> 정렬. 다만 더블클릭(열 너비 자동맞춤)과 겹치지 않도록
        약간 지연시켜 실행하고, 더블클릭이 감지되면 이 예약을 취소한다.
        선택(체크박스) 칸은 별도로 처리한다.
        """
        region = self.tree.identify_region(event.x, event.y)
        col = self._column_at(event.x)
        if col is None:
            return

        if col == CHECK_COL:
            if region == "cell":
                row_id = self.tree.identify_row(event.y)
                if row_id:
                    self._toggle_check(row_id)
            elif region == "heading":
                self._toggle_all_checks()
            return

        if region != "heading":
            return

        if not self._is_on_heading_text(col, event.x):
            return

        if self._sort_after_id:
            self.root.after_cancel(self._sort_after_id)
        self._sort_after_id = self.root.after(
            300, lambda c=col: self._trigger_sort(c)
        )

    def _is_on_heading_text(self, col, x):
        """정렬은 헤더 칸 전체가 아니라 글자가 그려진 부분을 정확히 클릭했을 때만 발생시킨다."""
        columns = self.tree["columns"]
        widths = [self.tree.column(c, "width") for c in columns]
        total_width = sum(widths)

        # 가로 스크롤바를 옮긴 상태에서는 컬럼들이 화면상 왼쪽으로 밀려있으므로,
        # 0에서부터 누적한 위치가 실제 화면 위치와 어긋난다. 현재 스크롤 위치(xview)만큼
        # 보정해야 실제로 보이는 글자 위치와 클릭 좌표가 맞는다.
        first_fraction, _ = self.tree.xview()
        scroll_offset = first_fraction * total_width

        x_start = -scroll_offset
        for c, width in zip(columns, widths):
            if c == col:
                break
            x_start += width
        col_width = self.tree.column(col, "width")

        text = self.tree.heading(col, "text")
        font = tkfont.Font(font=("맑은 고딕", 9, "bold"))
        text_width = font.measure(text)

        margin = 4
        text_start = x_start + max((col_width - text_width) // 2, 0) - margin
        text_end = text_start + text_width + margin * 2
        return text_start <= x <= text_end

    def _trigger_sort(self, col):
        self._sort_after_id = None
        self._sort_by(col)

    def _toggle_check(self, row_id):
        idx = self.tree.index(row_id)
        if not (0 <= idx < len(self.current_results)):
            return
        row = self.current_results[idx]
        row["_checked"] = not row.get("_checked", True)
        self.tree.set(row_id, CHECK_COL, CHECK_MARK if row["_checked"] else UNCHECK_MARK)
        self._update_check_header()

    def _toggle_all_checks(self):
        if not self.current_results:
            return
        all_checked = all(r.get("_checked", True) for r in self.current_results)
        new_state = not all_checked
        for row in self.current_results:
            row["_checked"] = new_state
        mark = CHECK_MARK if new_state else UNCHECK_MARK
        for iid in self.tree.get_children():
            self.tree.set(iid, CHECK_COL, mark)
        self._update_check_header()

    def _on_tree_double_click(self, event):
        if self._sort_after_id:
            self.root.after_cancel(self._sort_after_id)
            self._sort_after_id = None

        region = self.tree.identify_region(event.x, event.y)
        if region == "separator":
            col = self._column_at(event.x)
            if col:
                self._autofit_column(col)
            return

        if region == "cell":
            col = self._column_at(event.x)
            if col != "비고":
                return
            row_id = self.tree.identify_row(event.y)
            if row_id:
                self._start_edit_remark(row_id)

    def _column_at(self, x):
        col_id = self.tree.identify_column(x)
        try:
            col_index = int(col_id.replace("#", "")) - 1
        except ValueError:
            return None
        columns = self.tree["columns"]
        if col_index < 0 or col_index >= len(columns):
            return None
        return columns[col_index]

    def _autofit_column(self, col):
        font = tkfont.nametofont("TkDefaultFont")
        heading_font = tkfont.Font(font=("맑은 고딕", 9, "bold"))
        max_width = heading_font.measure(col) + 24
        for iid in self.tree.get_children():
            text = self.tree.set(iid, col)
            width = font.measure(text) + 24
            if width > max_width:
                max_width = width
        self.tree.column(col, width=max_width)
        self._redraw_grid_lines()

    # ---------- 비고란 직접 수정 ----------
    def _start_edit_remark(self, row_id):
        self._commit_edit()
        bbox = self.tree.bbox(row_id, "비고")
        if not bbox:
            return
        x, y, w, h = bbox
        value = self.tree.set(row_id, "비고")

        entry = tk.Entry(self.result_content)
        entry.insert(0, value)
        entry.select_range(0, "end")
        entry.place(x=x, y=y, width=w, height=h)
        entry.focus_set()
        entry.bind("<Return>", lambda e: self._commit_edit())
        entry.bind("<FocusOut>", lambda e: self._commit_edit())
        entry.bind("<Escape>", lambda e: self._cancel_edit())

        self._edit_entry = entry
        self._edit_row_id = row_id

    def _commit_edit(self):
        if self._edit_entry is None:
            return
        new_value = self._edit_entry.get()
        row_id = self._edit_row_id
        self._edit_entry.destroy()
        self._edit_entry = None
        self._edit_row_id = None

        if not self.tree.exists(row_id):
            return
        self.tree.set(row_id, "비고", new_value)
        idx = self.tree.index(row_id)
        if 0 <= idx < len(self.current_results):
            self.current_results[idx]["비고"] = new_value

    def _cancel_edit(self):
        if self._edit_entry is None:
            return
        self._edit_entry.destroy()
        self._edit_entry = None
        self._edit_row_id = None

    # ---------- 칸 구분선(그리드) ----------
    def _on_xscroll_set(self, *args):
        self._hsb.set(*args)
        # 가로 스크롤은 지연 없이 즉시 다시 그려야 선이 스크롤과 같이 매끄럽게 움직인다.
        self._redraw_grid_lines()

    def _schedule_grid_redraw(self, event=None):
        if self._grid_redraw_after_id:
            self.tree.after_cancel(self._grid_redraw_after_id)
        self._grid_redraw_after_id = self.tree.after(50, self._redraw_grid_lines)

    def _redraw_grid_lines(self):
        self._grid_redraw_after_id = None
        for w in self._grid_lines:
            w.destroy()
        self._grid_lines = []

        columns = self.tree["columns"]
        if not columns or not self.current_results:
            return

        self.tree.update_idletasks()
        children = self.tree.get_children()
        if not children:
            return

        tree_w = self.tree.winfo_width()
        tree_h = self.tree.winfo_height()
        if tree_w <= 1 or tree_h <= 1:
            return

        first_bbox = self.tree.bbox(children[0])
        if first_bbox:
            self._header_height = first_bbox[1]

        first_visible = self.tree.identify_row(self._header_height + 1)
        if not first_visible:
            first_visible = children[0]
        last_visible = self.tree.identify_row(tree_h - 1)
        if not last_visible:
            last_visible = children[-1]

        start_idx = self.tree.index(first_visible)
        end_idx = self.tree.index(last_visible)

        line_color = "#000000"

        bottom_bbox = self.tree.bbox(last_visible)
        if not bottom_bbox:
            return
        # 세로선은 헤더(표시할 항목 이름이 보이는 초록 줄)까지 이어지도록 맨 위(0)에서 시작한다.
        top = 0
        # 마지막 행이 트리 영역 경계에 걸쳐 일부만 보이는 경우, bbox는 잘리지 않은 전체
        # 행 높이를 반환하므로 그대로 쓰면 세로선이 스크롤바 영역까지 내려간다.
        # 실제 보이는 영역(tree_h)을 넘지 않도록 자른다.
        bottom = min(bottom_bbox[1] + bottom_bbox[3], tree_h)
        if bottom <= top:
            return

        for col in columns:
            bbox = self.tree.bbox(first_visible, col)
            if not bbox:
                continue
            x_right = bbox[0] + bbox[2]
            if x_right <= 0 or x_right >= tree_w:
                continue
            line = tk.Frame(self.result_content, background=line_color)
            line.place(x=x_right, y=top, width=1, height=bottom - top)
            self._grid_lines.append(line)

        for idx in range(start_idx, end_idx + 1):
            item = children[idx]
            bbox = self.tree.bbox(item)
            if not bbox:
                continue
            y_bottom = bbox[1] + bbox[3]
            if y_bottom > tree_h:
                continue
            line = tk.Frame(self.result_content, background=line_color)
            line.place(x=0, y=y_bottom, width=tree_w, height=1)
            self._grid_lines.append(line)

    # ---------- 내보내기 / 복사 ----------
    def _ask_export_scope(self):
        """반환값: 'all' | 'checked' | None(취소)"""
        dialog = tk.Toplevel(self.root)
        dialog.title("내보내기 범위")
        dialog.configure(background=CARD_BG)
        dialog.transient(self.root)
        dialog.resizable(False, False)

        result = {"choice": None}

        ttk.Label(
            dialog,
            text="전체 결과를 내보낼까요, 체크(☑)한 항목만 내보낼까요?",
            padding=16,
        ).pack()

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=(0, 12))

        def choose(value):
            result["choice"] = value
            dialog.destroy()

        ttk.Button(
            btn_frame, text="전체", command=lambda: choose("all"), style="Accent.TButton"
        ).pack(side="left", padx=6)
        ttk.Button(
            btn_frame, text="체크한 것만", command=lambda: choose("checked")
        ).pack(side="left", padx=6)
        ttk.Button(btn_frame, text="취소", command=lambda: choose(None)).pack(
            side="left", padx=6
        )

        dialog.protocol("WM_DELETE_WINDOW", lambda: choose(None))

        self.root.update_idletasks()
        x = self.root.winfo_rootx() + 250
        y = self.root.winfo_rooty() + 250
        dialog.geometry(f"+{x}+{y}")

        dialog.grab_set()
        dialog.wait_window()
        return result["choice"]

    def _rows_for_scope(self, scope):
        if scope == "checked":
            return [r for r in self.current_results if r.get("_checked", True)]
        return list(self.current_results)

    def _default_export_filename(self):
        keyword = self.keyword_var.get().strip()
        if not keyword:
            return "검색결과.xlsx"
        name = re.sub(r'[\\/:*?"<>|]', "_", keyword)
        name = re.sub(r"\s*,\s*", "_", name).strip()
        return f"{name}.xlsx" if name else "검색결과.xlsx"

    def _export_excel(self):
        if not self.current_results:
            messagebox.showwarning("알림", "내보낼 검색 결과가 없습니다.")
            return

        scope = self._ask_export_scope()
        if scope is None:
            return
        rows = self._rows_for_scope(scope)
        if not rows:
            messagebox.showwarning("알림", "내보낼 항목이 없습니다.")
            return

        path = filedialog.asksaveasfilename(
            title="엑셀로 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=self._default_export_filename(),
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "검색결과"

            header_fill = PatternFill("solid", fgColor=HEADER_BG.lstrip("#"))
            header_font = Font(color=HEADER_FG.lstrip("#"), bold=True)
            odd_fill = PatternFill("solid", fgColor="F5F6F7")
            thin = Side(style="thin", color="B0B0B0")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center_cols = {"보험코드", "약가", "요율"}

            ws.append(self.current_columns)
            for c in range(1, len(self.current_columns) + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for r, row in enumerate(rows, start=2):
                ws.append(
                    [format_value(f, row.get(f)) for f in self.current_columns]
                )
                color = row.get("_row_color")
                for c, col_name in enumerate(self.current_columns, start=1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal="center" if col_name in center_cols else "left",
                        vertical="center",
                    )
                    if r % 2 == 1:
                        cell.fill = odd_fill
                    if color:
                        cell.font = Font(color=color.lstrip("#"))

            for c, col_name in enumerate(self.current_columns, start=1):
                width = 34 if col_name in ("제품명", "성분", "비고") else 15
                ws.column_dimensions[get_column_letter(c)].width = width

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            ws.row_dimensions[1].height = 20

            wb.save(path)
            messagebox.showinfo("완료", f"저장했습니다.\n{path}")
        except Exception as e:
            messagebox.showerror("오류", f"저장 중 오류가 발생했습니다.\n{e}")

    def _copy_clipboard(self):
        if not self.current_results:
            messagebox.showwarning("알림", "복사할 검색 결과가 없습니다.")
            return

        scope = self._ask_export_scope()
        if scope is None:
            return
        rows = self._rows_for_scope(scope)
        if not rows:
            messagebox.showwarning("알림", "복사할 항목이 없습니다.")
            return

        lines = ["\t".join(self.current_columns)]
        for row in rows:
            lines.append(
                "\t".join(format_value(f, row.get(f)) for f in self.current_columns)
            )
        text = "\n".join(lines)

        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.status_var.set(f"클립보드에 {len(rows)}건 복사 완료")

    def _capture_copy(self):
        if not self.current_results:
            messagebox.showwarning("알림", "캡쳐할 검색 결과가 없습니다.")
            return

        scope = self._ask_export_scope()
        if scope is None:
            return
        rows = self._rows_for_scope(scope)
        if not rows:
            messagebox.showwarning("알림", "캡쳐할 항목이 없습니다.")
            return

        try:
            from PIL import ImageGrab

            if scope == "checked":
                self._refresh_rows(rows, self.current_columns)
                self.root.update()

            self.root.lift()
            self.root.update()
            # 범위선택 다이얼로그가 destroy()된 직후에는 Windows 화면이
            # 아직 다시 그려지기 전이라, 곧바로 캡쳐하면 다이얼로그 잔상이
            # 함께 찍힌다. 화면 갱신이 따라올 시간을 준 뒤 캡쳐한다.
            time.sleep(0.15)
            self.root.update()
            x = self.tree.winfo_rootx()
            y = self.tree.winfo_rooty()
            w = self.tree.winfo_width()
            h = self.tree.winfo_height()

            # 위젯 크기가 아니라 실제 데이터가 차지하는 영역만 캡쳐한다.
            # (트리 위젯은 창 크기에 맞춰 늘어나므로, 행이 적으면 아래쪽에
            # 빈 여백이 그대로 캡쳐되는 문제가 있었다.)
            columns = self.tree["columns"]
            if columns:
                content_w = sum(self.tree.column(c, "width") for c in columns)
                if content_w > 0:
                    w = min(content_w, w)

            children = self.tree.get_children()
            if children:
                last_item = children[min(len(rows), len(children)) - 1]
                last_bbox = self.tree.bbox(last_item)
                if last_bbox:
                    content_h = last_bbox[1] + last_bbox[3]
                    h = min(content_h, h)

            img = ImageGrab.grab(bbox=(x, y, x + w, y + h))
            self._image_to_clipboard(img)
            self.status_var.set("표를 이미지로 클립보드에 복사했습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"캡쳐 중 오류가 발생했습니다.\n{e}")
        finally:
            if scope == "checked":
                self._refresh_rows(self.current_results, self.current_columns)

    @staticmethod
    def _image_to_clipboard(image):
        import io
        import win32clipboard

        output = io.BytesIO()
        image.convert("RGB").save(output, "BMP")
        data = output.getvalue()[14:]  # BMP 헤더(14바이트) 제거 -> CF_DIB
        output.close()

        win32clipboard.OpenClipboard()
        try:
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
        finally:
            win32clipboard.CloseClipboard()


def main():
    root = tk.Tk()
    app = RateSheetApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
